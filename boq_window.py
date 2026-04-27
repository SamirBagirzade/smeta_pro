"""Smeta window implementation."""

import os
from datetime import timezone
import math
import re

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QTableWidget, QTableWidgetItem, QPushButton, QHeaderView, QMessageBox,
    QDialog, QSpinBox, QDialogButtonBox, QRadioButton, QButtonGroup,
    QFormLayout, QInputDialog, QComboBox, QDoubleSpinBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QShortcut, QKeySequence

from dialogs import SmetaItemDialog
from template_management import TemplateManagementWindow
from currency_settings import CurrencySettingsManager


class SmetaWindow(QMainWindow):
    """Bill of Quantities Window"""

    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.parent_window = parent
        self.db = db
        self.currency_manager = CurrencySettingsManager(self.db)
        self.boq_items = []  # List to store Smeta items
        self.next_id = 1
        self.boq_name = "Smeta 1"  # Default name
        self.string_count = 0
        self._updating_table = False
        self._breaker_ratings = [
            6, 10, 16, 20, 25, 32, 40, 50, 63,
            80, 100, 125, 160, 200, 250, 320, 400
        ]
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("📋 Bill of Quantities (Smeta)")
        self.setGeometry(150, 150, 1200, 650)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()

        # Title and Name section with Up/Down buttons
        title_layout = QHBoxLayout()

        title = QLabel("Bill of Quantities")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #2196F3; padding: 10px;")

        # Move Up/Down buttons (small, at top)
        self.move_up_btn = QPushButton("⬆️")
        self.move_up_btn.clicked.connect(self.move_item_up)
        self.move_up_btn.setFixedSize(40, 30)
        self.move_up_btn.setStyleSheet("background-color: #607D8B; color: white; border: none; border-radius: 4px; font-weight: bold;")
        self.move_up_btn.setToolTip("Yuxarı daşı (Ctrl+Up)")

        self.move_down_btn = QPushButton("⬇️")
        self.move_down_btn.clicked.connect(self.move_item_down)
        self.move_down_btn.setFixedSize(40, 30)
        self.move_down_btn.setStyleSheet("background-color: #607D8B; color: white; border: none; border-radius: 4px; font-weight: bold;")
        self.move_down_btn.setToolTip("Aşağı daşı (Ctrl+Down)")

        # Smeta Name input
        name_label = QLabel("Smeta Adı:")
        name_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        self.boq_name_input = QLineEdit(self.boq_name)
        self.boq_name_input.setMaximumWidth(200)
        self.boq_name_input.setStyleSheet("font-size: 14px; padding: 5px;")
        self.boq_name_input.textChanged.connect(self.update_boq_name)

        # String count input
        string_label = QLabel("String Sayı:")
        string_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        self.string_input = QSpinBox()
        self.string_input.setRange(0, 999999)
        self.string_input.setValue(self.string_count)
        self.string_input.setMaximumWidth(120)
        self.string_input.setStyleSheet("font-size: 14px; padding: 5px;")
        self.string_input.valueChanged.connect(self.update_string_count)

        # AC Breaker Wizard button
        self.ac_breaker_btn = QPushButton("⚡ AC Avtomat Sehirbazı")
        self.ac_breaker_btn.clicked.connect(self.open_ac_breaker_wizard)
        self.ac_breaker_btn.setStyleSheet("""
            QPushButton {
                background-color: #3F51B5;
                color: white;
                padding: 6px 12px;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #303F9F;
            }
        """)

        # AC Cable Wizard button
        self.ac_cable_btn = QPushButton("🔌 AC Kabel Sehirbazı")
        self.ac_cable_btn.clicked.connect(self.open_ac_cable_wizard)
        self.ac_cable_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 6px 12px;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)

        title_layout.addWidget(title)
        title_layout.addWidget(self.ac_breaker_btn)
        title_layout.addWidget(self.ac_cable_btn)
        title_layout.addWidget(self.move_up_btn)
        title_layout.addWidget(self.move_down_btn)
        title_layout.addStretch()
        title_layout.addWidget(name_label)
        title_layout.addWidget(self.boq_name_input)
        title_layout.addWidget(string_label)
        title_layout.addWidget(self.string_input)

        main_layout.addLayout(title_layout)

        # Table with margin column
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "№", "Adı", "Kateqoriya", "Miqdar", "Ölçü Vahidi", "Vahid Qiymət", "Cəmi", "Marja %", "Yekun", "Mənbə", "Qeyd", "Növ"
        ])

        # Table styling
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(
            QTableWidget.EditTrigger.SelectedClicked | QTableWidget.EditTrigger.EditKeyPressed
        )
        self.table.setSortingEnabled(True)  # Enable column sorting
        self.table.itemChanged.connect(self.on_table_item_changed)

        # Resize columns
        header = self.table.horizontalHeader()
        header.setSortIndicatorShown(True)
        header.sectionClicked.connect(self.sort_by_column)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # №
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Adı
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Kateqoriya
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Miqdar
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Ölçü Vahidi
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Vahid Qiymət
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Cəmi
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)  # Marja %
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)  # Yekun
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)  # Mənbə
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)  # Qeyd
        header.setSectionResizeMode(11, QHeaderView.ResizeMode.ResizeToContents)  # Növ

        main_layout.addWidget(self.table)
        self.table.cellDoubleClicked.connect(self.on_table_double_clicked)
        self.table.installEventFilter(self)

        # Summary labels
        summary_layout = QHBoxLayout()
        summary_layout.addStretch()

        self.cost_label = QLabel("Maya Dəyəri: 0.00 AZN")
        self.cost_label.setStyleSheet("font-size: 14px; color: #666; padding: 10px;")

        self.margin_total_label = QLabel("Ümumi Marja: 0.00 AZN")
        self.margin_total_label.setStyleSheet("font-size: 14px; color: #FF9800; padding: 10px;")

        self.summary_label = QLabel("Yekun Məbləğ: 0.00 AZN")
        self.summary_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #4CAF50; padding: 10px;")

        summary_layout.addWidget(self.cost_label)
        summary_layout.addWidget(self.margin_total_label)
        summary_layout.addWidget(self.summary_label)

        main_layout.addLayout(summary_layout)

        # Buttons - reorganized into two rows
        button_layout1 = QHBoxLayout()
        button_layout2 = QHBoxLayout()

        self.add_custom_btn = QPushButton("➕ Xüsusi Qeyd")
        self.add_custom_btn.clicked.connect(self.add_custom_item)
        self.add_custom_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px 15px;
                border: none;
                border-radius: 5px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)

        self.edit_btn = QPushButton("✏️ Redaktə Et")
        self.edit_btn.clicked.connect(self.edit_item)
        self.edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e68900;
            }
        """)

        self.delete_btn = QPushButton("🗑️ Sil")
        self.delete_btn.clicked.connect(self.delete_item)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)

        self.save_boq_btn = QPushButton("💾 Smeta Yadda Saxla")
        self.save_boq_btn.clicked.connect(self.save_boq)
        self.save_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)

        self.load_boq_btn = QPushButton("📂 Smeta Yüklə")
        self.load_boq_btn.clicked.connect(self.load_boq)
        self.load_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e68900;
            }
        """)

        self.export_excel_btn = QPushButton("📊 Excel-ə İxrac Et")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        self.load_cloud_boq_btn = QPushButton("☁️ Buluddan Yüklə")
        self.load_cloud_boq_btn.clicked.connect(self.load_from_cloud)
        self.load_cloud_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #00BCD4;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0097A7;
            }
        """)

        self.combine_boq_btn = QPushButton("🔗 Smeta-ları Birləşdir")
        self.combine_boq_btn.clicked.connect(self.combine_boqs_to_excel)
        self.combine_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #673AB7;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5E35B1;
            }
        """)

        # Template Management button
        self.template_mgmt_btn = QPushButton("📋 Şablonlar")
        self.template_mgmt_btn.clicked.connect(self.open_template_management)
        self.template_mgmt_btn.setStyleSheet("""
            QPushButton {
                background-color: #795548;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #6D4C41;
            }
        """)

        # Row 1: Edit actions
        button_layout1.addWidget(self.add_custom_btn)
        button_layout1.addWidget(self.edit_btn)
        button_layout1.addWidget(self.delete_btn)
        button_layout1.addWidget(self.save_boq_btn)
        button_layout1.addWidget(self.load_boq_btn)
        button_layout1.addStretch()

        # Row 2: Export/Cloud/Templates
        button_layout2.addWidget(self.load_cloud_boq_btn)
        button_layout2.addWidget(self.export_excel_btn)
        button_layout2.addWidget(self.combine_boq_btn)
        button_layout2.addWidget(self.template_mgmt_btn)
        button_layout2.addStretch()

        main_layout.addLayout(button_layout1)
        main_layout.addLayout(button_layout2)
        central_widget.setLayout(main_layout)

        # Apply light mode styling
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
        """)

        # Setup keyboard shortcuts
        self.setup_shortcuts()

    def setup_shortcuts(self):
        """Setup keyboard shortcuts for SmetaWindow"""
        # Ctrl+S: Save Smeta
        QShortcut(QKeySequence("Ctrl+S"), self).activated.connect(self.save_boq)

        # Ctrl+O: Load Smeta
        QShortcut(QKeySequence("Ctrl+O"), self).activated.connect(self.load_boq)

        # Ctrl+E: Edit selected item
        QShortcut(QKeySequence("Ctrl+E"), self).activated.connect(self.edit_item)
        # Enter handling is implemented via eventFilter on the table.

        # Delete: Delete selected item(s)
        QShortcut(QKeySequence("Delete"), self).activated.connect(self.delete_item)

        # Ctrl+Up: Move item up
        QShortcut(QKeySequence("Ctrl+Up"), self).activated.connect(self.move_item_up)

        # Ctrl+Down: Move item down
        QShortcut(QKeySequence("Ctrl+Down"), self).activated.connect(self.move_item_down)

    def add_from_database(self):
        """Add item from database"""
        if not self.db:
            QMessageBox.warning(self, "Xəta", "Verilənlər bazasına qoşulmayıbsınız!")
            return

        dialog = SmetaItemDialog(self, self.db, mode="add_from_db", string_count=self.string_count)
        if dialog.exec():
            data = dialog.get_data()
            data['id'] = self.next_id
            self.next_id += 1
            self.boq_items.append(data)
            self.refresh_table()

    def add_custom_item(self):
        """Add custom item (not from database)"""
        dialog = SmetaItemDialog(self, self.db, mode="custom", string_count=self.string_count)
        if dialog.exec():
            data = dialog.get_data()
            data['id'] = self.next_id
            self.next_id += 1
            self.boq_items.append(data)
            self.refresh_table()

    def open_ac_breaker_wizard(self):
        """Collect inverter specs, calculate breaker ratings, and add to BoQ."""
        dialog = QDialog(self)
        dialog.setWindowTitle("AC Avtomat Sehirbazı")
        layout = QVBoxLayout()

        form_layout = QFormLayout()
        phase_layout = QHBoxLayout()

        single_phase_radio = QRadioButton("Tək faza (220V)")
        three_phase_radio = QRadioButton("3 faza (380V)")
        single_phase_radio.setChecked(True)

        phase_group = QButtonGroup(dialog)
        phase_group.addButton(single_phase_radio, 1)
        phase_group.addButton(three_phase_radio, 3)

        phase_layout.addWidget(single_phase_radio)
        phase_layout.addWidget(three_phase_radio)
        phase_layout.addStretch()

        inverter_spin = QSpinBox()
        inverter_spin.setRange(1, 9999)
        inverter_spin.setValue(1)
        inverter_spin.setMaximumWidth(120)

        form_layout.addRow("Faza:", phase_layout)
        form_layout.addRow("İnverter sayı:", inverter_spin)
        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.setLayout(layout)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        phase_count = phase_group.checkedId()
        voltage = 220 if phase_count == 1 else 380
        inverter_count = inverter_spin.value()

        needed_counts = {}
        overflow_used = False
        for idx in range(inverter_count):
            kw_value, ok = QInputDialog.getDouble(
                self,
                f"İnverter {idx + 1} kVt",
                "kVt daxil edin:",
                decimals=2,
                min=0.0,
                max=1e9
            )
            if not ok:
                return
            watts = kw_value * 1000.0
            breaker_rating, overflow = self._calculate_breaker_rating(watts, voltage)
            overflow_used = overflow_used or overflow
            needed_counts[breaker_rating] = needed_counts.get(breaker_rating, 0) + 1

        if overflow_used:
            QMessageBox.information(
                self,
                "Məlumat",
                "Bəzi invertrlər üçün hesablanan cərəyan maksimum avtomat dəyərini keçdi. "
                f"{self._breaker_ratings[-1]}A istifadə edildi."
            )

        for rating, count in sorted(needed_counts.items()):
            name = f"AC Breaker C{phase_count}x{rating} A"
            existing_item = self._find_boq_item_by_name(name)
            if existing_item:
                existing_item['quantity'] = existing_item.get('quantity', 0) + count
                existing_item['total'] = existing_item.get('quantity', 0) * existing_item.get('unit_price_azn', 0)
            else:
                data = {
                    'product_id': None,
                    'name': name,
                    'quantity': float(count),
                    'unit': "Ədəd",
                    'unit_price': 0.0,
                    'unit_price_azn': 0.0,
                    'total': 0.0,
                    'currency': "AZN",
                    'margin_percent': 0.0,
                    'quantity_round': True,
                    'price_round': False,
                    'is_custom': True,
                    'category': "Elektrik;AC Mühafizə",
                    'source': "",
                    'note': ""
                }
                data['id'] = self.next_id
                self.next_id += 1
                self.boq_items.append(data)

        self.refresh_table()

    def open_ac_cable_wizard(self):
        """AC Cable sizing wizard based on IEC standards."""
        dialog = QDialog(self)
        dialog.setWindowTitle("AC Kabel Sehirbazı")
        layout = QVBoxLayout()

        form_layout = QFormLayout()

        # Current input
        current_spin = QDoubleSpinBox()
        current_spin.setRange(0.1, 10000.0)
        current_spin.setValue(10.0)
        current_spin.setSuffix(" A")
        form_layout.addRow("Cərəyan (A):", current_spin)

        # Voltage input
        voltage_spin = QSpinBox()
        voltage_spin.setRange(100, 1000)
        voltage_spin.setValue(400)
        voltage_spin.setSuffix(" V")
        form_layout.addRow("Gərginlik (V):", voltage_spin)

        # Distance input
        distance_spin = QDoubleSpinBox()
        distance_spin.setRange(0.1, 10000.0)
        distance_spin.setValue(100.0)
        distance_spin.setSuffix(" m")
        form_layout.addRow("Məsafə (m):", distance_spin)

        # Voltage drop %
        drop_spin = QDoubleSpinBox()
        drop_spin.setRange(0.1, 10.0)
        drop_spin.setValue(3.0)
        drop_spin.setSuffix(" %")
        form_layout.addRow("İcazə verilən gərginlik düşümü (%):", drop_spin)

        # Material
        material_combo = QComboBox()
        material_combo.addItems(["Copper", "Aluminum"])
        form_layout.addRow("Material:", material_combo)

        # Insulation
        insulation_combo = QComboBox()
        insulation_combo.addItems(["PVC", "XLPE"])
        form_layout.addRow("İzoliyasiya:", insulation_combo)

        # Installation
        install_combo = QComboBox()
        install_combo.addItems(["Open air", "In Unperforated tray", "In duct underground"])
        form_layout.addRow("Quraşdırma növü:", install_combo)

        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.setLayout(layout)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        # Get values
        current = current_spin.value()
        voltage = voltage_spin.value()
        distance = distance_spin.value()
        max_drop_percent = drop_spin.value()
        material = material_combo.currentText()
        insulation = insulation_combo.currentText()
        installation = install_combo.currentText()

        # Calculate cable size
        cable_size = self._calculate_cable_size(current, voltage, distance, max_drop_percent, material, insulation, installation)

        if cable_size:
            # Add to BOQ
            name = f"AC Cable {cable_size} mm² {material} {insulation} ({installation})"
            existing_item = self._find_boq_item_by_name(name)
            if existing_item:
                existing_item['quantity'] = existing_item.get('quantity', 0) + distance / 1000.0  # km
                existing_item['total'] = existing_item.get('quantity', 0) * existing_item.get('unit_price_azn', 0)
            else:
                data = {
                    'product_id': None,
                    'name': name,
                    'quantity': distance / 1000.0,  # km
                    'unit': "km",
                    'unit_price': 0.0,
                    'unit_price_azn': 0.0,
                    'total': 0.0,
                    'currency': "AZN",
                    'margin_percent': 0.0,
                    'quantity_round': False,
                    'price_round': False,
                    'is_custom': True,
                    'category': "Elektrik;Kabel",
                    'source': "",
                    'note': f"Calculated for {current}A, {voltage}V, {distance}m, {max_drop_percent}% drop"
                }
                data['id'] = self.next_id
                self.next_id += 1
                self.boq_items.append(data)

            self.refresh_table()
            QMessageBox.information(self, "Uğur", f"Kabel ölçüsü hesablandı: {cable_size} mm²")
        else:
            QMessageBox.warning(self, "Xəta", "Uyğun kabel ölçüsü tapılmadı.")

    def _calculate_cable_size(self, current, voltage, distance, max_drop_percent, material, insulation, installation):
        """Calculate minimum cable size based on IEC standards."""
        # Resistivity (ohm.mm²/m)
        rho = 0.0175 if material == "Copper" else 0.028  # Copper: 0.0175, Aluminum: 0.028

        # Derating factors for installation
        derating = {
            "Open air": 1.0,
            "In Unperforated tray": 0.9,
            "In duct underground": 0.8
        }.get(installation, 1.0)

        # Ampacity table (simplified, based on IEC for PVC/XLPE)
        ampacity_pvc = {
            1.5: 17, 2.5: 24, 4: 32, 6: 41, 10: 57, 16: 76, 25: 101, 35: 125, 50: 151, 70: 192, 95: 232, 120: 269, 150: 302, 185: 341, 240: 400, 300: 464
        }
        ampacity_xlpe = {
            1.5: 20, 2.5: 27, 4: 36, 6: 46, 10: 64, 16: 85, 25: 114, 35: 141, 50: 170, 70: 216, 95: 261, 120: 293, 150: 339, 185: 382, 240: 445, 300: 517
        }
        ampacity = ampacity_xlpe if insulation == "XLPE" else ampacity_pvc

        # Standard sizes
        sizes = sorted(ampacity.keys())

        for size in sizes:
            # Check ampacity
            max_current = ampacity[size] * derating
            if current > max_current:
                continue

            # Check voltage drop
            # R = rho / size (ohm/m)
            r_per_m = rho / size
            # For AC, voltage drop approx V_drop = I * R * L * 2 (round trip)
            v_drop = current * r_per_m * distance * 2  # Volts
            drop_percent = (v_drop / voltage) * 100
            if drop_percent <= max_drop_percent:
                return size

        return None

    def _calculate_breaker_rating(self, watts, voltage):
        """Return the next standard breaker rating with 15% upsize."""
        if voltage == 0:
            current = 0.0
        elif voltage == 380:
            current = watts / (voltage * math.sqrt(3))
        else:
            current = watts / voltage
        needed = current * 1.15
        for rating in self._breaker_ratings:
            if needed <= rating:
                return rating, False
        return self._breaker_ratings[-1], True

    def _find_boq_item_by_name(self, name):
        for item in self.boq_items:
            if item.get('name') == name:
                return item
        return None

    def edit_item(self):
        """Edit selected item"""
        if not self.isActiveWindow():
            return
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Redaktə etmək üçün qeyd seçin!")
            return

        item = self.boq_items[selected_row]
        mode = "custom" if item.get('is_custom') else "edit"

        dialog = SmetaItemDialog(self, self.db, item=item, mode=mode, string_count=self.string_count)
        if dialog.exec():
            data = dialog.get_data()
            data['id'] = item['id']
            self.boq_items[selected_row] = data
            self.refresh_table()

    def on_table_double_clicked(self, row, column):
        """Handle double click on table cells."""
        if column in (3, 7):
            return
        self.edit_item()

    def eventFilter(self, source, event):
        if source is self.table and event.type() == event.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                self.edit_item()
                return True
        return super().eventFilter(source, event)

    def delete_item(self):
        """Delete selected item(s)"""
        # Get all selected rows
        selected_rows = self.table.selectionModel().selectedRows()

        if not selected_rows:
            QMessageBox.warning(self, "Xəbərdarlıq", "Silmək üçün qeyd seçin!")
            return

        # Confirm deletion
        if len(selected_rows) == 1:
            message = "Bu qeydi silmək istədiyinizdən əminsiniz?"
        else:
            message = f"{len(selected_rows)} qeydi silmək istədiyinizdən əminsiniz?"

        reply = QMessageBox.question(
            self,
            "Təsdiq",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Get row indices and sort in descending order
            # This allows us to delete from bottom to top without affecting other indices
            rows_to_delete = sorted([index.row() for index in selected_rows], reverse=True)

            # Delete items from the list
            for row in rows_to_delete:
                del self.boq_items[row]

            self.refresh_table()

    def move_item_up(self):
        """Move selected item up in the list"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Yuxarı daşımaq üçün qeyd seçin!")
            return

        if selected_row == 0:
            return  # Already at top

        # Swap items
        self.boq_items[selected_row], self.boq_items[selected_row - 1] = \
            self.boq_items[selected_row - 1], self.boq_items[selected_row]

        # Renumber items
        self._renumber_items()
        self.refresh_table()

        # Keep selection on the moved item
        self.table.selectRow(selected_row - 1)

    def move_item_down(self):
        """Move selected item down in the list"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Aşağı daşımaq üçün qeyd seçin!")
            return

        if selected_row >= len(self.boq_items) - 1:
            return  # Already at bottom

        # Swap items
        self.boq_items[selected_row], self.boq_items[selected_row + 1] = \
            self.boq_items[selected_row + 1], self.boq_items[selected_row]

        # Renumber items
        self._renumber_items()
        self.refresh_table()

        # Keep selection on the moved item
        self.table.selectRow(selected_row + 1)

    def _renumber_items(self):
        """Renumber all items sequentially"""
        for idx, item in enumerate(self.boq_items, start=1):
            item['id'] = idx
        self.next_id = len(self.boq_items) + 1

    def refresh_table(self):
        """Refresh the Smeta table"""
        self._updating_table = True
        self.table.setRowCount(0)

        for item in self.boq_items:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            # Get margin percent (default 0)
            margin_pct = item.get('margin_percent', 0)
            currency = item.get('currency', 'AZN') or 'AZN'
            unit_price = item.get('unit_price', 0)
            unit_price_azn = item.get('unit_price_azn')
            if unit_price_azn is None:
                unit_price_azn = self.currency_manager.convert_to_azn(unit_price, currency)
                item['unit_price_azn'] = unit_price_azn

            cost_total = item.get('total')
            if cost_total is None:
                cost_total = item.get('quantity', 0) * unit_price_azn
                item['total'] = cost_total
            final_total = cost_total * (1 + margin_pct / 100)

            # Column 0: №
            id_item = QTableWidgetItem(str(item['id']))
            id_item.setFlags(id_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 0, id_item)
            # Column 1: Adı
            name_item = QTableWidgetItem(item['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 1, name_item)
            # Column 2: Kateqoriya
            category_item = QTableWidgetItem(item.get('category', '') or 'N/A')
            category_item.setFlags(category_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 2, category_item)
            # Column 3: Miqdar
            quantity_item = QTableWidgetItem(f"{item['quantity']:.2f}")
            quantity_item.setFlags(quantity_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 3, quantity_item)
            # Column 4: Ölçü Vahidi
            unit_item = QTableWidgetItem(item['unit'])
            unit_item.setFlags(unit_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 4, unit_item)
            # Column 5: Vahid Qiymət
            if currency == "AZN":
                price_text = f"{unit_price:.2f} AZN"
            else:
                price_text = f"AZN {unit_price_azn:.2f} ({unit_price:.2f} {currency})"
            price_item = QTableWidgetItem(price_text)
            price_item.setFlags(price_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 5, price_item)
            # Column 6: Cəmi (cost)
            cost_item = QTableWidgetItem(f"{cost_total:.2f}")
            cost_item.setFlags(cost_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 6, cost_item)
            # Column 7: Marja %
            margin_item = QTableWidgetItem(f"{margin_pct:.1f}%")
            margin_item.setFlags(margin_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 7, margin_item)
            # Column 8: Yekun (with margin)
            final_item = QTableWidgetItem(f"{final_total:.2f}")
            final_item.setFlags(final_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 8, final_item)
            # Column 9: Mənbə
            source_item = QTableWidgetItem(item.get('source', '') or 'N/A')
            source_item.setFlags(source_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 9, source_item)
            # Column 10: Qeyd
            note_item = QTableWidgetItem(item.get('note', '') or 'N/A')
            note_item.setFlags(note_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 10, note_item)
            # Column 11: Növ
            item_type = "Xüsusi" if item.get('is_custom') else "DB"
            type_item = QTableWidgetItem(item_type)
            type_item.setFlags(type_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_position, 11, type_item)

        # Update summary
        self.update_summary()
        self._updating_table = False

    def on_table_item_changed(self, item):
        """Handle inline edits in the Smeta table."""
        if self._updating_table:
            return
        if item.column() not in (3, 7):
            return

        row = item.row()
        if row < 0 or row >= len(self.boq_items):
            return

        if item.column() == 3:
            text = item.text().strip().replace(',', '.')
            if not text:
                quantity = self.boq_items[row].get('quantity', 1)
            else:
                try:
                    quantity = float(text)
                except ValueError:
                    quantity = self.boq_items[row].get('quantity', 1)

            quantity = max(0.01, quantity)
            self.boq_items[row]['quantity'] = quantity
            unit_price_azn = self.boq_items[row].get('unit_price_azn')
            if unit_price_azn is None:
                currency = self.boq_items[row].get('currency', 'AZN') or 'AZN'
                unit_price = self.boq_items[row].get('unit_price', 0)
                unit_price_azn = self.currency_manager.convert_to_azn(unit_price, currency)
                self.boq_items[row]['unit_price_azn'] = unit_price_azn
            total = quantity * unit_price_azn
            self.boq_items[row]['total'] = total

            margin_pct = self.boq_items[row].get('margin_percent', 0)
            final_total = total * (1 + margin_pct / 100)

            self._updating_table = True
            item.setText(f"{quantity:.2f}")
            cost_item = self.table.item(row, 6)
            if cost_item:
                cost_item.setText(f"{total:.2f}")
            final_item = self.table.item(row, 8)
            if final_item:
                final_item.setText(f"{final_total:.2f}")
            self._updating_table = False

            self.update_summary()
            return
        if item.column() == 5:
            text = item.text().strip().replace(',', '.')
            numbers = re.findall(r"[-+]?\d*\.?\d+", text)
            if not numbers:
                unit_price_azn = self.boq_items[row].get('unit_price_azn', 0)
            else:
                try:
                    unit_price_azn = float(numbers[0])
                except ValueError:
                    unit_price_azn = self.boq_items[row].get('unit_price_azn', 0)

            unit_price_azn = max(0.0, unit_price_azn)
            self.boq_items[row]['unit_price_azn'] = unit_price_azn
            self.boq_items[row]['unit_price'] = unit_price_azn
            self.boq_items[row]['currency'] = 'AZN'

            quantity = self.boq_items[row].get('quantity', 1)
            total = quantity * unit_price_azn
            self.boq_items[row]['total'] = total

            margin_pct = self.boq_items[row].get('margin_percent', 0)
            final_total = total * (1 + margin_pct / 100)

            self._updating_table = True
            item.setText(f"{unit_price_azn:.2f} AZN")
            cost_item = self.table.item(row, 6)
            if cost_item:
                cost_item.setText(f"{total:.2f}")
            final_item = self.table.item(row, 8)
            if final_item:
                final_item.setText(f"{final_total:.2f}")
            self._updating_table = False

            self.update_summary()
            return

        text = item.text().strip().replace('%', '')
        text = text.replace(',', '.')
        if not text:
            margin_pct = 0.0
        else:
            try:
                margin_pct = float(text)
            except ValueError:
                margin_pct = self.boq_items[row].get('margin_percent', 0)

        margin_pct = max(0.0, min(100.0, margin_pct))
        self.boq_items[row]['margin_percent'] = margin_pct

        cost_total = self.boq_items[row].get('total', 0)
        final_total = cost_total * (1 + margin_pct / 100)

        self._updating_table = True
        item.setText(f"{margin_pct:.1f}%")
        final_item = self.table.item(row, 8)
        if final_item:
            final_item.setText(f"{final_total:.2f}")
        self._updating_table = False

        self.update_summary()

    def update_summary(self):
        """Update the summary labels with cost total, margin total, and final amount"""
        cost_total = sum(item['total'] for item in self.boq_items)
        margin_total = sum(item['total'] * (item.get('margin_percent', 0) / 100) for item in self.boq_items)
        final_total = cost_total + margin_total

        self.cost_label.setText(f"Maya Dəyəri: {cost_total:.2f} AZN")
        self.margin_total_label.setText(f"Ümumi Marja: {margin_total:.2f} AZN")
        self.summary_label.setText(f"Yekun Məbləğ: {final_total:.2f} AZN")

    def sort_by_column(self, column):
        """Sort boq_items by the clicked column"""
        # Map column indices to item keys
        column_keys = {
            0: 'id',
            1: 'name',
            2: 'category',
            3: 'quantity',
            4: 'unit',
            5: 'unit_price',
            6: 'total',
            7: 'margin_percent',
            8: 'final_total',  # computed, not stored
            9: 'source',
            10: 'note',
            11: 'is_custom'
        }

        key = column_keys.get(column)
        if not key:
            return

        # Check current sort order from header
        header = self.table.horizontalHeader()
        current_order = header.sortIndicatorOrder()

        # Sort the boq_items list
        reverse = (current_order == Qt.SortOrder.DescendingOrder)

        try:
            if key == 'final_total':
                # Computed field: total * (1 + margin_percent/100)
                self.boq_items.sort(key=lambda x: x.get('total', 0) * (1 + x.get('margin_percent', 0) / 100), reverse=reverse)
            elif key in ['quantity', 'unit_price', 'total', 'id', 'margin_percent']:
                # Numeric sort
                if key == 'unit_price':
                    self.boq_items.sort(key=lambda x: float(x.get('unit_price_azn', x.get('unit_price', 0)) or 0), reverse=reverse)
                else:
                    self.boq_items.sort(key=lambda x: float(x.get(key, 0) or 0), reverse=reverse)
            elif key == 'is_custom':
                # Boolean sort
                self.boq_items.sort(key=lambda x: x.get(key, False), reverse=reverse)
            else:
                # String sort
                self.boq_items.sort(key=lambda x: str(x.get(key, '') or '').lower(), reverse=reverse)

            # Renumber after sort
            self._renumber_items()

            # Refresh table without triggering sort again
            self.table.setSortingEnabled(False)
            self.refresh_table()
            self.table.setSortingEnabled(True)

        except Exception as e:
            print(f"Sort error: {e}")

    def export_to_excel(self):
        """Export Smeta to Excel file"""
        if not self.boq_items:
            QMessageBox.warning(self, "Xəbərdarlıq", "Smeta boşdur! İxrac etmək üçün məhsul əlavə edin.")
            return

        try:
            # Import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                QMessageBox.critical(
                    self,
                    "Xəta",
                    "openpyxl kitabxanası tapılmadı!\n\nYükləmək üçün terminal-da:\npip install openpyxl"
                )
                return

            # Ask user for file location
            from PyQt6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Smeta-u Excel-ə İxrac Et",
                "Smeta.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bill of Quantities"

            # Define styles
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF", size=12)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add title
            ws.merge_cells('A1:L1')
            ws['A1'] = "BILL OF QUANTITIES (BOQ)"
            ws['A1'].font = Font(bold=True, size=16, color="2196F3")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Margin header style
            margin_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")

            # Add headers
            headers = ["№", "Adı", "Kateqoriya", "Miqdar", "Ölçü Vahidi", "Vahid Qiymət (AZN)", "Cəmi (AZN)", "Marja %", "Yekun (AZN)", "Mənbə", "Qeyd", "Növ"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                # Use orange color for margin columns
                if col_num in [8, 9]:
                    cell.fill = margin_fill
                else:
                    cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Add data
            data_start_row = 4
            data_end_row = data_start_row + len(self.boq_items) - 1
            for row_num, item in enumerate(self.boq_items, data_start_row):
                margin_pct = item.get('margin_percent', 0)

                # Column 1: №
                ws.cell(row=row_num, column=1, value=item['id']).border = border
                # Column 2: Adı
                ws.cell(row=row_num, column=2, value=item['name']).border = border
                # Column 3: Kateqoriya
                ws.cell(row=row_num, column=3, value=item.get('category', '') or 'N/A').border = border
                # Column 4: Miqdar
                ws.cell(row=row_num, column=4, value=item['quantity']).border = border
                # Column 5: Ölçü Vahidi
                ws.cell(row=row_num, column=5, value=item['unit']).border = border
                # Column 6: Vahid Qiymət (AZN)
                unit_price_azn = item.get('unit_price_azn')
                if unit_price_azn is None:
                    currency = item.get('currency', 'AZN') or 'AZN'
                    unit_price_azn = self.currency_manager.convert_to_azn(item.get('unit_price', 0), currency)
                ws.cell(row=row_num, column=6, value=unit_price_azn).border = border
                # Column 7: Cəmi (cost)
                ws.cell(row=row_num, column=7, value=f"=D{row_num}*F{row_num}").border = border
                # Column 8: Marja %
                ws.cell(row=row_num, column=8, value=margin_pct).border = border
                # Column 9: Yekun (with margin)
                ws.cell(row=row_num, column=9, value=f"=G{row_num}*(1+H{row_num}/100)").border = border
                # Column 10: Mənbə
                ws.cell(row=row_num, column=10, value=item.get('source', '') or 'N/A').border = border
                # Column 11: Qeyd
                ws.cell(row=row_num, column=11, value=item.get('note', '') or 'N/A').border = border
                # Column 12: Növ
                item_type = "Xüsusi" if item.get('is_custom') else "DB"
                ws.cell(row=row_num, column=12, value=item_type).border = border

                # Format numbers
                ws.cell(row=row_num, column=4).number_format = '0.00'
                ws.cell(row=row_num, column=6).number_format = '0.00'
                ws.cell(row=row_num, column=7).number_format = '0.00'
                ws.cell(row=row_num, column=8).number_format = '0.0'
                ws.cell(row=row_num, column=9).number_format = '0.00'

            # Add total row
            total_row = len(self.boq_items) + 5

            # Cost total label
            ws.merge_cells(f'A{total_row}:F{total_row}')
            cost_label_cell = ws[f'A{total_row}']
            cost_label_cell.value = "MAYA DƏYƏRİ:"
            cost_label_cell.fill = header_fill
            cost_label_cell.font = total_font
            cost_label_cell.alignment = Alignment(horizontal="right", vertical="center")
            cost_label_cell.border = border

            # Cost total value
            cost_value_cell = ws[f'G{total_row}']
            cost_value_cell.value = f"=SUM(G{data_start_row}:G{data_end_row})"
            cost_value_cell.fill = header_fill
            cost_value_cell.font = total_font
            cost_value_cell.alignment = Alignment(horizontal="center", vertical="center")
            cost_value_cell.border = border
            cost_value_cell.number_format = '0.00'

            # Margin total label
            margin_label_cell = ws[f'H{total_row}']
            margin_label_cell.value = "TOPLAM:"
            margin_label_cell.fill = margin_fill
            margin_label_cell.font = total_font
            margin_label_cell.alignment = Alignment(horizontal="center", vertical="center")
            margin_label_cell.border = border

            # Final total value
            final_value_cell = ws[f'I{total_row}']
            final_value_cell.value = f"=SUM(I{data_start_row}:I{data_end_row})"
            final_value_cell.fill = total_fill
            final_value_cell.font = total_font
            final_value_cell.alignment = Alignment(horizontal="center", vertical="center")
            final_value_cell.border = border
            final_value_cell.number_format = '0.00'

            # Fill remaining cells in total row
            for col in ['J', 'K', 'L']:
                ws[f'{col}{total_row}'].fill = total_fill
                ws[f'{col}{total_row}'].border = border

            # Add margin summary row
            margin_row = total_row + 1
            ws.merge_cells(f'A{margin_row}:F{margin_row}')
            margin_summary_cell = ws[f'A{margin_row}']
            margin_summary_cell.value = "MARJA:"
            margin_summary_cell.fill = margin_fill
            margin_summary_cell.font = total_font
            margin_summary_cell.alignment = Alignment(horizontal="right", vertical="center")
            margin_summary_cell.border = border

            margin_summary_value = ws[f'G{margin_row}']
            margin_summary_value.value = f"=I{total_row}-G{total_row}"
            margin_summary_value.fill = margin_fill
            margin_summary_value.font = total_font
            margin_summary_value.alignment = Alignment(horizontal="center", vertical="center")
            margin_summary_value.border = border
            margin_summary_value.number_format = '0.00'

            # Fill remaining cells
            for col in ['H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{margin_row}'].fill = margin_fill
                ws[f'{col}{margin_row}'].border = border

            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 10

            # Category totals sheet
            category_ws = wb.create_sheet(title="Category Totals")
            category_ws["A1"] = "Kateqoriya"
            category_ws["B1"] = "Yekun (AZN)"
            for cell in ("A1", "B1"):
                category_ws[cell].fill = header_fill
                category_ws[cell].font = header_font
                category_ws[cell].alignment = header_alignment
                category_ws[cell].border = border

            categories = []
            seen = set()
            for item in self.boq_items:
                category = item.get("category", "") or "N/A"
                if category not in seen:
                    seen.add(category)
                    categories.append(category)

            for idx, category in enumerate(categories, start=2):
                category_ws.cell(row=idx, column=1, value=category).border = border
                category_ws.cell(
                    row=idx,
                    column=2,
                    value=f"=SUMIF('Bill of Quantities'!$C${data_start_row}:$C${data_end_row},A{idx},'Bill of Quantities'!$I${data_start_row}:$I${data_end_row})"
                ).border = border
                category_ws.cell(row=idx, column=2).number_format = '0.00'

            total_row = len(categories) + 2
            category_ws.cell(row=total_row, column=1, value="CƏMİ").border = border
            category_ws.cell(row=total_row, column=1).font = total_font
            category_ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})").border = border
            category_ws.cell(row=total_row, column=2).font = total_font
            category_ws.cell(row=total_row, column=2).number_format = '0.00'

            category_ws.column_dimensions['A'].width = 25
            category_ws.column_dimensions['B'].width = 18

            # Save file
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Uğurlu",
                f"Smeta uğurla Excel-ə ixrac edildi!\n\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"İxrac zamanı xəta:\n{str(e)}")

    def update_boq_name(self):
        """Update Smeta name from input field"""
        self.boq_name = self.boq_name_input.text().strip() or "Smeta 1"

    def update_string_count(self):
        """Update string count from input field"""
        self.string_count = int(self.string_input.value())

    def save_boq(self):
        """Save Smeta to JSON file with optional cloud save"""
        if not self.boq_items:
            QMessageBox.warning(self, "Xəbərdarlıq", "Smeta boşdur! Yadda saxlamaq üçün məhsul əlavə edin.")
            return

        try:
            import json
            from PyQt6.QtWidgets import QFileDialog, QCheckBox

            # Create custom dialog with checkbox
            dialog = QDialog(self)
            dialog.setWindowTitle("Smeta-u Yadda Saxla")
            dialog.setMinimumWidth(400)

            layout = QVBoxLayout()

            # Info label
            info_label = QLabel("Smeta-u harada saxlamaq istəyirsiniz?")
            layout.addWidget(info_label)

            # Cloud save checkbox
            cloud_checkbox = QCheckBox("☁️ Buludda da saxla (MongoDB)")
            cloud_checkbox.setChecked(True)  # Default checked
            cloud_checkbox.setStyleSheet("padding: 10px; font-size: 12px;")
            layout.addWidget(cloud_checkbox)

            # Buttons
            button_layout = QHBoxLayout()
            save_btn = QPushButton("💾 Yadda Saxla")
            save_btn.setDefault(True)
            cancel_btn = QPushButton("❌ Ləğv Et")

            save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)

            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
            """)

            button_layout.addWidget(save_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)

            dialog.setLayout(layout)

            # Connect buttons
            save_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            # Show dialog
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            save_to_cloud = cloud_checkbox.isChecked()

            # Ask user for file location
            default_name = f"{self.boq_name}.json" if self.boq_name else "Smeta.json"

            # Set default directory to saved_boqs folder
            saved_boqs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "saved_boqs")
            os.makedirs(saved_boqs_dir, exist_ok=True)
            default_path = os.path.join(saved_boqs_dir, default_name)

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Smeta-u Yadda Saxla",
                default_path,
                "JSON Files (*.json)"
            )

            if not file_path:
                return

            # Prepare data for saving
            save_data = {
                'boq_name': self.boq_name,
                'next_id': self.next_id,
                'string_count': self.string_count,
                'items': self.boq_items
            }

            # Save to local file
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)

            success_message = f"Smeta lokal faylda yadda saxlanıldı:\n{file_path}"

            # Save to cloud if checkbox is checked
            if save_to_cloud and self.db:
                try:
                    _, is_new = self.db.save_boq_to_cloud(
                        self.boq_name,
                        self.boq_items,
                        self.next_id,
                        self.string_count
                    )
                    if is_new:
                        success_message += "\n\n✅ Buludda da saxlanıldı (yeni)!"
                    else:
                        success_message += "\n\n✅ Buludda yeniləndi!"
                except Exception as cloud_error:
                    success_message += f"\n\n⚠️ Bulud saxlama xətası: {cloud_error}"

            QMessageBox.information(
                self,
                "Uğurlu",
                success_message
            )

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Smeta yadda saxlanarkən xəta:\n{str(e)}")

    def load_boq(self):
        """Load Smeta from JSON file and update prices from database"""
        try:
            import json
            from PyQt6.QtWidgets import QFileDialog

            # Ask user for file to load
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Smeta Yüklə",
                "",
                "JSON Files (*.json)"
            )

            if not file_path:
                return

            # Confirm if current Smeta will be replaced
            if self.boq_items:
                reply = QMessageBox.question(
                    self,
                    "Təsdiq",
                    "Mövcud Smeta məlumatları əvəz olunacaq. Davam etmək istəyirsiniz?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return

            # Load from file
            with open(file_path, 'r', encoding='utf-8') as f:
                save_data = json.load(f)

            self.boq_name = save_data.get('boq_name', 'Smeta 1')
            self.boq_name_input.setText(self.boq_name)
            self.next_id = save_data.get('next_id', 1)
            self.string_count = int(save_data.get('string_count', 0))
            self.string_input.setValue(self.string_count)
            loaded_items = save_data.get('items', [])

            # Update prices from database for items that came from DB
            updated_count = 0
            for item in loaded_items:
                if not item.get('is_custom') and item.get('product_id') and self.db:
                    try:
                        # Get current product data from database
                        product = self.db.read_product(item['product_id'])
                        if product:
                            # Update price, category, source, and note from database
                            old_price = item['unit_price']
                            new_price = float(product['price']) if product.get('price') else 0

                            item['unit_price'] = new_price
                            currency = product.get('currency', 'AZN') or 'AZN'
                            item['currency'] = currency
                            item['unit_price_azn'] = product.get('price_azn', new_price)
                            item['total'] = item['quantity'] * item['unit_price_azn']
                            item['category'] = product.get('category', '') or ''
                            item['source'] = product.get('mehsul_menbeyi', '') or ''
                            item['note'] = product.get('qeyd', '') or ''

                            if old_price != new_price:
                                updated_count += 1
                    except Exception:
                        # If product not found or error, keep the saved data
                        pass
                currency = item.get('currency', 'AZN') or 'AZN'
                if 'unit_price_azn' not in item or item.get('unit_price_azn') is None:
                    unit_price = item.get('unit_price', 0)
                    item['unit_price_azn'] = self.currency_manager.convert_to_azn(unit_price, currency)
                if 'total' not in item or item.get('total') is None:
                    item['total'] = item.get('quantity', 0) * item.get('unit_price_azn', 0)

            self.boq_items = loaded_items
            self.refresh_table()

            message = f"Smeta uğurla yükləndi!\n\n{len(loaded_items)} məhsul yükləndi."
            if updated_count > 0:
                message += f"\n{updated_count} məhsulun qiyməti yeniləndi."

            QMessageBox.information(self, "Uğurlu", message)

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Smeta yüklənərkən xəta:\n{str(e)}")

    def load_from_cloud(self):
        """Load Smeta from cloud (MongoDB)"""
        if not self.db:
            QMessageBox.warning(self, "Xəbərdarlıq", "Verilənlər bazasına qoşulun!")
            return

        try:
            # Get all cloud Smetas
            cloud_boqs = self.db.get_all_cloud_boqs()

            if not cloud_boqs:
                QMessageBox.information(self, "Məlumat", "Buludda heç bir Smeta tapılmadı!")
                return

            # Create selection dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Buluddan Smeta Yüklə")
            dialog.setMinimumWidth(600)
            dialog.setMinimumHeight(500)

            layout = QVBoxLayout()

            # Title
            title = QLabel("Yükləmək üçün Smeta seçin:")
            title.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px;")
            layout.addWidget(title)

            # Search input
            search_layout = QHBoxLayout()
            search_input = QLineEdit()
            search_input.setPlaceholderText("🔍 Smeta adına görə axtar...")
            search_input.setStyleSheet("""
                QLineEdit {
                    padding: 8px 12px;
                    border: 2px solid #ddd;
                    border-radius: 5px;
                    font-size: 13px;
                }
                QLineEdit:focus {
                    border-color: #2196F3;
                }
            """)
            search_layout.addWidget(search_input)
            layout.addLayout(search_layout)

            # List widget for Smetas
            from PyQt6.QtWidgets import QListWidget, QListWidgetItem
            boq_list = QListWidget()
            boq_list.setStyleSheet("""
                QListWidget {
                    font-size: 13px;
                    padding: 5px;
                }
                QListWidget::item {
                    padding: 10px;
                    border-bottom: 1px solid #ddd;
                }
                QListWidget::item:hover {
                    background-color: #e3f2fd;
                }
                QListWidget::item:selected {
                    background-color: #2196F3;
                    color: white;
                }
            """)

            # Store all Smetas for filtering
            all_boqs = cloud_boqs

            def format_boq_item(boq):
                """Format Smeta data for display"""
                item_count = len(boq.get('items', []))
                updated_at = boq.get('updated_at', '')
                if updated_at:
                    try:
                        # Convert UTC to local time
                        if updated_at.tzinfo is None:
                            updated_at = updated_at.replace(tzinfo=timezone.utc)
                        local_time = updated_at.astimezone()
                        updated_str = local_time.strftime('%Y-%m-%d %H:%M')
                    except Exception:
                        updated_str = str(updated_at)
                else:
                    updated_str = "Naməlum"
                return f"{boq['name']} ({item_count} məhsul) - Son yenilənmə: {updated_str}"

            def populate_list(boqs):
                """Populate the list with Smetas"""
                boq_list.clear()
                for boq in boqs:
                    item_text = format_boq_item(boq)
                    list_item = QListWidgetItem(item_text)
                    list_item.setData(Qt.ItemDataRole.UserRole, boq['id'])
                    boq_list.addItem(list_item)

            def search_boqs():
                """Search Smetas based on input"""
                search_term = search_input.text().strip()
                if not search_term:
                    populate_list(all_boqs)
                else:
                    try:
                        # Search from database
                        filtered_boqs = self.db.search_cloud_boqs(search_term)
                        populate_list(filtered_boqs)
                    except Exception:
                        # Fallback to local filtering
                        filtered_boqs = [b for b in all_boqs if search_term.lower() in b['name'].lower()]
                        populate_list(filtered_boqs)

            # Connect search input to search function
            search_input.textChanged.connect(search_boqs)

            # Initial population
            populate_list(all_boqs)

            layout.addWidget(boq_list)

            # Buttons
            button_layout = QHBoxLayout()

            load_btn = QPushButton("📥 Yüklə")
            load_btn.setDefault(True)
            delete_btn = QPushButton("🗑️ Sil")
            cancel_btn = QPushButton("❌ Ləğv Et")

            load_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)

            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
            """)

            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #9E9E9E;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #757575;
                }
            """)

            button_layout.addWidget(load_btn)
            button_layout.addWidget(delete_btn)
            button_layout.addStretch()
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)

            dialog.setLayout(layout)

            # Handle load button
            def load_selected():
                selected_items = boq_list.selectedItems()
                if not selected_items:
                    QMessageBox.warning(dialog, "Xəbərdarlıq", "Smeta seçin!")
                    return

                boq_id = selected_items[0].data(Qt.ItemDataRole.UserRole)

                # Confirm if current Smeta will be replaced
                if self.boq_items:
                    reply = QMessageBox.question(
                        dialog,
                        "Təsdiq",
                        "Mövcud Smeta məlumatları əvəz olunacaq. Davam etmək istəyirsiniz?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    if reply != QMessageBox.StandardButton.Yes:
                        return

                # Load Smeta from cloud
                boq_data = self.db.load_boq_from_cloud(boq_id)
                if boq_data:
                    self.boq_name = boq_data.get('name', 'Smeta 1')
                    self.boq_name_input.setText(self.boq_name)
                    self.next_id = boq_data.get('next_id', 1)
                    self.string_count = int(boq_data.get('string_count', 0))
                    self.string_input.setValue(self.string_count)
                    loaded_items = boq_data.get('items', [])

                    # Update prices from database for items that came from DB
                    updated_count = 0
                    for item in loaded_items:
                        if not item.get('is_custom') and item.get('product_id') and self.db:
                            try:
                                product = self.db.read_product(item['product_id'])
                                if product:
                                    old_price = item['unit_price']
                                    new_price = float(product['price']) if product.get('price') else 0

                                    item['unit_price'] = new_price
                                    currency = product.get('currency', 'AZN') or 'AZN'
                                    item['currency'] = currency
                                    item['unit_price_azn'] = product.get('price_azn', new_price)
                                    item['total'] = item['quantity'] * item['unit_price_azn']
                                    item['category'] = product.get('category', '') or ''
                                    item['source'] = product.get('mehsul_menbeyi', '') or ''
                                    item['note'] = product.get('qeyd', '') or ''

                                    if old_price != new_price:
                                        updated_count += 1
                            except Exception:
                                pass
                        currency = item.get('currency', 'AZN') or 'AZN'
                        if 'unit_price_azn' not in item or item.get('unit_price_azn') is None:
                            unit_price = item.get('unit_price', 0)
                            item['unit_price_azn'] = self.currency_manager.convert_to_azn(unit_price, currency)
                        if 'total' not in item or item.get('total') is None:
                            item['total'] = item.get('quantity', 0) * item.get('unit_price_azn', 0)

                    self.boq_items = loaded_items
                    self.refresh_table()

                    message = f"Smeta buluddan yükləndi!\n\n{len(loaded_items)} məhsul yükləndi."
                    if updated_count > 0:
                        message += f"\n{updated_count} məhsulun qiyməti yeniləndi."

                    QMessageBox.information(self, "Uğurlu", message)
                    dialog.accept()
                else:
                    QMessageBox.critical(dialog, "Xəta", "Smeta yüklənə bilmədi!")

            # Handle delete button
            def delete_selected():
                selected_items = boq_list.selectedItems()
                if not selected_items:
                    QMessageBox.warning(dialog, "Xəbərdarlıq", "Smeta seçin!")
                    return

                boq_id = selected_items[0].data(Qt.ItemDataRole.UserRole)
                boq_name = selected_items[0].text().split(' (')[0]

                reply = QMessageBox.question(
                    dialog,
                    "Təsdiq",
                    f"'{boq_name}' Smeta-nu buluddan silmək istədiyinizdən əminsiniz?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    if self.db.delete_cloud_boq(boq_id):
                        boq_list.takeItem(boq_list.row(selected_items[0]))
                        QMessageBox.information(dialog, "Uğurlu", "Smeta buluddan silindi!")

                        if boq_list.count() == 0:
                            QMessageBox.information(dialog, "Məlumat", "Buludda daha Smeta qalmadı.")
                            dialog.accept()
                    else:
                        QMessageBox.critical(dialog, "Xəta", "Smeta silinə bilmədi!")

            load_btn.clicked.connect(load_selected)
            delete_btn.clicked.connect(delete_selected)
            cancel_btn.clicked.connect(dialog.reject)

            # Double click to load
            boq_list.itemDoubleClicked.connect(load_selected)

            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Buluddan yükləmə xətası:\n{str(e)}")

    def combine_boqs_to_excel(self):
        """Combine multiple Smetas into a single Excel file"""
        try:
            import json
            from PyQt6.QtWidgets import QFileDialog
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            # Ask user to select multiple Smeta files
            file_paths, _ = QFileDialog.getOpenFileNames(
                self,
                "Smeta Fayllarını Seçin",
                "",
                "JSON Files (*.json)"
            )

            if not file_paths or len(file_paths) < 2:
                QMessageBox.warning(self, "Xəbərdarlıq", "Ən azı 2 Smeta faylı seçin!")
                return

            # Load all Smetas
            boqs = []
            for file_path in file_paths:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        boq_data = json.load(f)
                        boqs.append({
                            'name': boq_data.get('boq_name', os.path.basename(file_path)),
                            'items': boq_data.get('items', [])
                        })
                except Exception as e:
                    QMessageBox.warning(self, "Xəbərdarlıq", f"Fayl oxuna bilmədi: {os.path.basename(file_path)}\n{str(e)}")

            if len(boqs) < 2:
                QMessageBox.warning(self, "Xəbərdarlıq", "Ən azı 2 Smeta uğurla yüklənməlidir!")
                return

            # Create a unified list of all unique items (by name)
            all_items_dict = {}
            item_order = []
            for boq in boqs:
                for item in boq['items']:
                    item_key = item['name']
                    if item_key not in all_items_dict:
                        currency = item.get('currency', 'AZN') or 'AZN'
                        unit_price = item.get('unit_price', 0)
                        unit_price_azn = item.get('unit_price_azn')
                        if unit_price_azn is None:
                            unit_price_azn = self.currency_manager.convert_to_azn(unit_price, currency)
                        all_items_dict[item_key] = {
                            'name': item['name'],
                            'unit': item.get('unit', 'ədəd'),
                            'unit_price': unit_price_azn,
                            'category': item.get('category', ''),
                            'source': item.get('source', ''),
                            'note': item.get('note', ''),
                            'boq_data': {}
                        }
                        item_order.append(item_key)

            # Fill quantities for each Smeta
            for boq_idx, boq in enumerate(boqs):
                boq_name = boq['name']
                for item in boq['items']:
                    item_key = item['name']
                    all_items_dict[item_key]['boq_data'][boq_name] = {
                        'quantity': item.get('quantity', 0),
                        'margin_percent': item.get('margin_percent', 0)
                    }

            # Ensure all items have entries for all Smetas (fill with 0 if missing)
            for item_data in all_items_dict.values():
                for boq in boqs:
                    if boq['name'] not in item_data['boq_data']:
                        item_data['boq_data'][boq['name']] = {
                            'quantity': 0,
                            'margin_percent': 0
                        }

            # Ask user for output file
            output_path, _ = QFileDialog.getSaveFileName(
                self,
                "Birləşdirilmiş Smeta-u Yadda Saxla",
                "Birləşdirilmiş_Smeta.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not output_path:
                return

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Combined Smeta"

            # Define styles
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF", size=11)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            num_boqs = len(boqs)
            last_col_idx = 5 + (2 * num_boqs) + 2
            ws.merge_cells(f'A1:{get_column_letter(last_col_idx)}1')
            ws['A1'] = "BİRLƏŞDİRİLMİŞ BILL OF QUANTITIES (BOQ)"
            ws['A1'].font = Font(bold=True, size=16, color="2196F3")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Headers
            headers = ["№", "Adı", "Kateqoriya", "Ölçü Vahidi", "Vahid Qiymət (AZN)"]
            for boq in boqs:
                headers.append(f"{boq['name']}\n(Miqdar)")
                headers.append(f"{boq['name']}\n(Yekun AZN)")
            headers.append("Cəmi\nMiqdar")
            headers.append("Cəmi\nYekun (AZN)")

            col_num = 1
            for header in headers:
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                col_num += 1

            # Data rows
            row_num = 4
            qty_cols = []
            final_cols = []
            base_col = 6
            for i in range(num_boqs):
                qty_cols.append(base_col + (i * 2))
                final_cols.append(base_col + (i * 2) + 1)
            total_qty_col_idx = base_col + (num_boqs * 2)
            total_final_col_idx = total_qty_col_idx + 1

            for idx, item_key in enumerate(item_order, 1):
                item_data = all_items_dict[item_key]
                ws.cell(row=row_num, column=1, value=idx).border = border
                ws.cell(row=row_num, column=2, value=item_data['name']).border = border
                ws.cell(row=row_num, column=3, value=item_data['category'] or 'N/A').border = border
                ws.cell(row=row_num, column=4, value=item_data['unit']).border = border
                ws.cell(row=row_num, column=5, value=item_data['unit_price']).border = border
                ws.cell(row=row_num, column=5).number_format = '0.00'

                # Quantities and margin totals for each Smeta
                col = 6
                unit_price_value = float(item_data.get('unit_price') or 0)
                for boq in boqs:
                    boq_entry = item_data['boq_data'][boq['name']]
                    qty = boq_entry['quantity']
                    margin_pct = boq_entry['margin_percent']
                    try:
                        margin_pct_value = float(margin_pct)
                    except (TypeError, ValueError):
                        margin_pct_value = 0.0
                    final_total = unit_price_value * float(qty or 0) * (1 + margin_pct_value / 100)

                    ws.cell(row=row_num, column=col, value=qty).border = border
                    ws.cell(row=row_num, column=col).number_format = '0.00'
                    col += 1

                    ws.cell(row=row_num, column=col, value=final_total).border = border
                    ws.cell(row=row_num, column=col).number_format = '#,##0.00'
                    col += 1

                # Total quantity column (SUM formula of qty columns)
                qty_sum_cells = ",".join([f"{get_column_letter(c)}{row_num}" for c in qty_cols])
                total_qty_cell = ws.cell(row=row_num, column=total_qty_col_idx)
                total_qty_cell.value = f"=SUM({qty_sum_cells})"
                total_qty_cell.border = border
                total_qty_cell.number_format = '0.00'
                total_qty_cell.font = Font(bold=True)

                # Total final price column (SUM of margin totals)
                final_sum_cells = ",".join([f"{get_column_letter(c)}{row_num}" for c in final_cols])
                total_final_cell = ws.cell(row=row_num, column=total_final_col_idx)
                total_final_cell.value = f"=SUM({final_sum_cells})"
                total_final_cell.border = border
                total_final_cell.number_format = '#,##0.00'
                total_final_cell.font = Font(bold=True)

                row_num += 1

            # Add grand total row
            total_row = row_num + 1
            ws.merge_cells(f'A{total_row}:{get_column_letter(total_final_col_idx - 1)}{total_row}')
            total_label_cell = ws[f'A{total_row}']
            total_label_cell.value = "ÜMUMİ MƏBLƏĞ:"
            total_label_cell.fill = total_fill
            total_label_cell.font = total_font
            total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_label_cell.border = border

            # Grand total formula
            total_final_col_letter = get_column_letter(total_final_col_idx)
            grand_total_cell = ws[f'{total_final_col_letter}{total_row}']
            grand_total_cell.value = f"=SUM({total_final_col_letter}4:{total_final_col_letter}{row_num - 1})"
            grand_total_cell.fill = total_fill
            grand_total_cell.font = total_font
            grand_total_cell.alignment = Alignment(horizontal="center", vertical="center")
            grand_total_cell.border = border
            grand_total_cell.number_format = '#,##0.00'

            # Adjust column widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18

            for i in range((2 * num_boqs) + 2):  # Smeta qty/final + Total Qty + Total Final
                col_letter = chr(70 + i)  # Start from F
                ws.column_dimensions[col_letter].width = 14

            # Save file
            wb.save(output_path)

            QMessageBox.information(
                self,
                "Uğurlu",
                f"Smeta-lar uğurla birləşdirildi!\n\n{len(boqs)} Smeta birləşdirildi\n{len(all_items_dict)} unikal məhsul\n\n{output_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Smeta-lar birləşdirilərkən xəta:\n{str(e)}")

    def open_template_management(self):
        """Open Template Management window"""
        if not self.db:
            QMessageBox.warning(self, "Xəbərdarlıq", "Verilənlər bazasına qoşulmayıbsınız!")
            return

        dialog = TemplateManagementWindow(self, self.db, self)
        dialog.exec()
