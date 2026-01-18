"""
PyQt6 CRUD Application for Product Management with MongoDB
Database Structure:
- _id: ObjectId (Primary Key, auto-generated)
- mehsulun_adi: string (Required)
- price: float
- mehsul_menbeyi: string
- qeyd: string
- olcu_vahidi: string
- category: string
- price_last_changed: datetime (timestamp of last price change)
- image_id: ObjectId (Optional, reference to GridFS file)

Images are stored in GridFS for efficient binary storage.
Double-click any product row to view its image.
"""

import sys
import re
import json
import os
from datetime import datetime, timezone
from urllib.parse import quote_plus
from pymongo import MongoClient, ASCENDING, TEXT
from bson.objectid import ObjectId
import gridfs
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QLineEdit, QLabel,
    QMessageBox, QDialog, QFormLayout, QDoubleSpinBox, QTextEdit,
    QGroupBox, QHeaderView, QSpinBox, QFileDialog, QScrollArea, QMenu
)
from PyQt6.QtCore import Qt, QTimer, QByteArray
from PyQt6.QtGui import QFont, QIcon, QPalette, QColor, QPixmap, QShortcut, QKeySequence


class DatabaseManager:
    """Handles all MongoDB database operations"""

    def __init__(self, host="", port=27017, database="admin",
                 username="admin", password=""):
        """
        Initialize database connection parameters

        Args:
            host: Database host
            port: Database port (default: 27017)
            database: Database name
            username: Database username (optional)
            password: Database password (optional)
        """
        self.host = host
        self.port = port
        self.database = database
        self.username = username
        self.password = password
        self.client = None
        self.db = None
        self.collection = None
        self.fs = None  # GridFS for storing images
        self.connect()
        self.setup_indexes()

    def connect(self):
        """Create and maintain database connection"""
        try:
            # Build connection string with URL-encoded credentials
            if self.username and self.password:
                # URL-encode username and password to handle special characters
                encoded_username = quote_plus(self.username)
                encoded_password = quote_plus(self.password)
                connection_string = f"mongodb://{encoded_username}:{encoded_password}@{self.host}:{self.port}/{self.database}"
            else:
                connection_string = f"mongodb://{self.host}:{self.port}/"

            self.client = MongoClient(
                connection_string,
                serverSelectionTimeoutMS=5000,
                connectTimeoutMS=5000
            )
            self.db = self.client[self.database]
            self.collection = self.db['products']
            self.boq_collection = self.db['boqs']  # Collection for cloud BoQ storage
            self.fs = gridfs.GridFS(self.db)  # Initialize GridFS for image storage

            # Test connection
            self.client.server_info()

        except Exception as e:
            raise Exception(f"Database connection error: {e}")

    def setup_indexes(self):
        """Create indexes for faster searching"""
        try:
            # Create text indexes for search functionality
            self.collection.create_index([
                ("mehsulun_adi", TEXT),
                ("mehsul_menbeyi", TEXT),
                ("qeyd", TEXT),
                ("category", TEXT)
            ])

            # Create regular indexes for common queries
            self.collection.create_index([("mehsulun_adi", ASCENDING)])
            self.collection.create_index([("category", ASCENDING)])

            # Indexes created successfully (silent for GUI app)
        except Exception:
            # Could not create indexes (silent for GUI app)
            pass

    def create_product(self, mehsulun_adi, price, mehsul_menbeyi, qeyd, olcu_vahidi, category, image_id=None):
        """Create a new product"""
        try:
            product = {
                'mehsulun_adi': mehsulun_adi,
                'price': price,
                'mehsul_menbeyi': mehsul_menbeyi,
                'qeyd': qeyd,
                'olcu_vahidi': olcu_vahidi,
                'category': category,
                'price_last_changed': datetime.now(timezone.utc)
            }
            if image_id:
                product['image_id'] = image_id
            result = self.collection.insert_one(product)
            return str(result.inserted_id)
        except Exception as e:
            raise Exception(f"Failed to create product: {e}")

    def read_all_products(self):
        """Read all products"""
        try:
            products = list(self.collection.find().sort("_id", ASCENDING))
            # Convert ObjectId to string for display
            for product in products:
                product['id'] = str(product['_id'])
            return products
        except Exception as e:
            raise Exception(f"Failed to read products: {e}")

    def read_product(self, product_id):
        """Read a single product by ID"""
        try:
            # Handle both string and ObjectId
            if isinstance(product_id, str):
                product_id = ObjectId(product_id)

            product = self.collection.find_one({'_id': product_id})
            if product:
                product['id'] = str(product['_id'])
            return product
        except Exception as e:
            raise Exception(f"Failed to read product: {e}")

    def get_price_history(self, product_id):
        """Get price history for a product"""
        try:
            if isinstance(product_id, str):
                product_id = ObjectId(product_id)

            product = self.collection.find_one({'_id': product_id})
            if product:
                return product.get('price_history', [])
            return []
        except Exception as e:
            raise Exception(f"Failed to get price history: {e}")

    def update_product(self, product_id, mehsulun_adi, price, mehsul_menbeyi, qeyd, olcu_vahidi, category, image_id=None):
        """Update an existing product"""
        try:
            # Handle both string and ObjectId
            if isinstance(product_id, str):
                product_id = ObjectId(product_id)

            # Get current product to check if price changed
            current_product = self.collection.find_one({'_id': product_id})

            update_data = {
                '$set': {
                    'mehsulun_adi': mehsulun_adi,
                    'price': price,
                    'mehsul_menbeyi': mehsul_menbeyi,
                    'qeyd': qeyd,
                    'olcu_vahidi': olcu_vahidi,
                    'category': category
                }
            }

            # If price changed, update the timestamp and add to history
            if current_product and current_product.get('price') != price:
                update_data['$set']['price_last_changed'] = datetime.now(timezone.utc)
                # Add to price history
                old_price = current_product.get('price', 0)
                history_entry = {
                    'old_price': old_price,
                    'new_price': price,
                    'changed_at': datetime.now(timezone.utc)
                }
                update_data['$push'] = {'price_history': history_entry}

            # Handle image update
            if image_id is not None:
                # Delete old image if exists
                if current_product and current_product.get('image_id'):
                    try:
                        self.fs.delete(current_product['image_id'])
                    except:
                        pass

                if image_id:  # New image uploaded
                    update_data['$set']['image_id'] = image_id
                else:  # Image removed
                    update_data['$unset'] = {'image_id': ''}

            result = self.collection.update_one({'_id': product_id}, update_data)
            return result.modified_count > 0 or result.matched_count > 0
        except Exception as e:
            raise Exception(f"Failed to update product: {e}")

    def delete_product(self, product_id):
        """Delete a product"""
        try:
            # Handle both string and ObjectId
            if isinstance(product_id, str):
                product_id = ObjectId(product_id)

            result = self.collection.delete_one({'_id': product_id})
            return result.deleted_count > 0
        except Exception as e:
            raise Exception(f"Failed to delete product: {e}")

    def search_products(self, search_term):
        """Search products by name, source, note, or category"""
        try:
            # Use text search for better performance
            products = list(self.collection.find(
                {'$text': {'$search': search_term}}
            ).sort("_id", ASCENDING))

            # If no results with text search, try regex (fallback)
            if not products:
                # Escape special regex characters to treat them as literals
                escaped_term = re.escape(search_term)
                regex_pattern = {'$regex': escaped_term, '$options': 'i'}
                products = list(self.collection.find({
                    '$or': [
                        {'mehsulun_adi': regex_pattern},
                        {'mehsul_menbeyi': regex_pattern},
                        {'qeyd': regex_pattern},
                        {'category': regex_pattern}
                    ]
                }).sort("_id", ASCENDING))

            # Convert ObjectId to string
            for product in products:
                product['id'] = str(product['_id'])

            return products
        except Exception as e:
            raise Exception(f"Failed to search products: {e}")

    def test_connection(self):
        """Test database connection"""
        try:
            info = self.client.server_info()
            return True, f"MongoDB version: {info.get('version', 'Unknown')}"
        except Exception as e:
            return False, str(e)

    def save_image(self, image_data, filename):
        """Save image to GridFS and return the file ID"""
        try:
            file_id = self.fs.put(image_data, filename=filename)
            return file_id
        except Exception as e:
            raise Exception(f"Failed to save image: {e}")

    def get_image(self, file_id):
        """Retrieve image from GridFS"""
        try:
            if isinstance(file_id, str):
                file_id = ObjectId(file_id)
            image_file = self.fs.get(file_id)
            return image_file.read()
        except Exception as e:
            raise Exception(f"Failed to retrieve image: {e}")

    def delete_image(self, file_id):
        """Delete image from GridFS"""
        try:
            if isinstance(file_id, str):
                file_id = ObjectId(file_id)
            self.fs.delete(file_id)
        except Exception as e:
            raise Exception(f"Failed to delete image: {e}")

    # BoQ Cloud Storage Methods
    def save_boq_to_cloud(self, boq_name, boq_items, next_id):
        """Save BoQ to MongoDB cloud"""
        try:
            boq_data = {
                'name': boq_name,
                'items': boq_items,
                'next_id': next_id,
                'created_at': datetime.now(timezone.utc),
                'updated_at': datetime.now(timezone.utc)
            }

            # Check if BoQ with same name exists
            existing_boq = self.boq_collection.find_one({'name': boq_name})

            if existing_boq:
                # Update existing BoQ
                boq_data['created_at'] = existing_boq.get('created_at', datetime.now(timezone.utc))
                boq_data['updated_at'] = datetime.now(timezone.utc)
                self.boq_collection.update_one(
                    {'name': boq_name},
                    {'$set': boq_data}
                )
                return str(existing_boq['_id']), False  # False = updated
            else:
                # Insert new BoQ
                result = self.boq_collection.insert_one(boq_data)
                return str(result.inserted_id), True  # True = created new

        except Exception as e:
            raise Exception(f"Failed to save BoQ to cloud: {e}")

    def get_all_cloud_boqs(self):
        """Get list of all BoQs from cloud"""
        try:
            boqs = list(self.boq_collection.find().sort("updated_at", -1))
            for boq in boqs:
                boq['id'] = str(boq['_id'])
            return boqs
        except Exception as e:
            raise Exception(f"Failed to retrieve cloud BoQs: {e}")

    def search_cloud_boqs(self, search_term):
        """Search BoQs by name"""
        try:
            escaped_term = re.escape(search_term)
            regex_pattern = {'$regex': escaped_term, '$options': 'i'}
            boqs = list(self.boq_collection.find(
                {'name': regex_pattern}
            ).sort("updated_at", -1))
            for boq in boqs:
                boq['id'] = str(boq['_id'])
            return boqs
        except Exception as e:
            raise Exception(f"Failed to search cloud BoQs: {e}")

    def load_boq_from_cloud(self, boq_id):
        """Load a specific BoQ from cloud"""
        try:
            if isinstance(boq_id, str):
                boq_id = ObjectId(boq_id)

            boq = self.boq_collection.find_one({'_id': boq_id})
            if boq:
                boq['id'] = str(boq['_id'])
            return boq
        except Exception as e:
            raise Exception(f"Failed to load BoQ from cloud: {e}")

    def delete_cloud_boq(self, boq_id):
        """Delete a BoQ from cloud"""
        try:
            if isinstance(boq_id, str):
                boq_id = ObjectId(boq_id)

            result = self.boq_collection.delete_one({'_id': boq_id})
            return result.deleted_count > 0
        except Exception as e:
            raise Exception(f"Failed to delete cloud BoQ: {e}")

    # BoQ Template Methods
    def save_template(self, template_name, items):
        """Save a BoQ template to cloud"""
        try:
            # Initialize template collection if not exists
            if not hasattr(self, 'template_collection'):
                self.template_collection = self.db['boq_templates']

            # Remove quantity-specific data for template
            template_items = []
            for item in items:
                template_item = {
                    'name': item['name'],
                    'unit': item.get('unit', ''),
                    'unit_price': item.get('unit_price', 0),
                    'category': item.get('category', ''),
                    'source': item.get('source', ''),
                    'note': item.get('note', ''),
                    'is_custom': item.get('is_custom', False),
                    'product_id': item.get('product_id')
                }
                template_items.append(template_item)

            # Check if template exists
            existing = self.template_collection.find_one({'name': template_name})
            if existing:
                # Update existing template
                self.template_collection.update_one(
                    {'name': template_name},
                    {'$set': {
                        'items': template_items,
                        'updated_at': datetime.now(timezone.utc)
                    }}
                )
            else:
                # Create new template
                self.template_collection.insert_one({
                    'name': template_name,
                    'items': template_items,
                    'created_at': datetime.now(timezone.utc),
                    'updated_at': datetime.now(timezone.utc)
                })
            return True
        except Exception as e:
            raise Exception(f"Failed to save template: {e}")

    def get_all_templates(self):
        """Get all BoQ templates"""
        try:
            if not hasattr(self, 'template_collection'):
                self.template_collection = self.db['boq_templates']

            templates = list(self.template_collection.find().sort("updated_at", -1))
            for t in templates:
                t['id'] = str(t['_id'])
            return templates
        except Exception as e:
            raise Exception(f"Failed to get templates: {e}")

    def load_template(self, template_id):
        """Load a specific template"""
        try:
            if not hasattr(self, 'template_collection'):
                self.template_collection = self.db['boq_templates']

            if isinstance(template_id, str):
                template_id = ObjectId(template_id)

            template = self.template_collection.find_one({'_id': template_id})
            if template:
                template['id'] = str(template['_id'])
            return template
        except Exception as e:
            raise Exception(f"Failed to load template: {e}")

    def delete_template(self, template_id):
        """Delete a template"""
        try:
            if not hasattr(self, 'template_collection'):
                self.template_collection = self.db['boq_templates']

            if isinstance(template_id, str):
                template_id = ObjectId(template_id)

            result = self.template_collection.delete_one({'_id': template_id})
            return result.deleted_count > 0
        except Exception as e:
            raise Exception(f"Failed to delete template: {e}")

    # Project Management Methods
    def _init_project_collection(self):
        """Initialize project collection if not exists"""
        if not hasattr(self, 'project_collection'):
            self.project_collection = self.db['projects']

    def create_project(self, name, description="", status="Aktiv"):
        """Create a new project"""
        try:
            self._init_project_collection()
            project = {
                'name': name,
                'description': description,
                'status': status,
                'boq_ids': [],
                'created_at': datetime.now(timezone.utc),
                'updated_at': datetime.now(timezone.utc)
            }
            result = self.project_collection.insert_one(project)
            return str(result.inserted_id)
        except Exception as e:
            raise Exception(f"Failed to create project: {e}")

    def get_all_projects(self):
        """Get all projects"""
        try:
            self._init_project_collection()
            projects = list(self.project_collection.find().sort("updated_at", -1))
            for p in projects:
                p['id'] = str(p['_id'])
            return projects
        except Exception as e:
            raise Exception(f"Failed to get projects: {e}")

    def get_project(self, project_id):
        """Get a specific project"""
        try:
            self._init_project_collection()
            if isinstance(project_id, str):
                project_id = ObjectId(project_id)
            project = self.project_collection.find_one({'_id': project_id})
            if project:
                project['id'] = str(project['_id'])
            return project
        except Exception as e:
            raise Exception(f"Failed to get project: {e}")

    def update_project(self, project_id, name=None, description=None, status=None, boq_ids=None):
        """Update a project"""
        try:
            self._init_project_collection()
            if isinstance(project_id, str):
                project_id = ObjectId(project_id)

            update_data = {'$set': {'updated_at': datetime.now(timezone.utc)}}
            if name is not None:
                update_data['$set']['name'] = name
            if description is not None:
                update_data['$set']['description'] = description
            if status is not None:
                update_data['$set']['status'] = status
            if boq_ids is not None:
                update_data['$set']['boq_ids'] = boq_ids

            result = self.project_collection.update_one({'_id': project_id}, update_data)
            return result.modified_count > 0 or result.matched_count > 0
        except Exception as e:
            raise Exception(f"Failed to update project: {e}")

    def delete_project(self, project_id):
        """Delete a project"""
        try:
            self._init_project_collection()
            if isinstance(project_id, str):
                project_id = ObjectId(project_id)
            result = self.project_collection.delete_one({'_id': project_id})
            return result.deleted_count > 0
        except Exception as e:
            raise Exception(f"Failed to delete project: {e}")

    def add_boq_to_project(self, project_id, boq_id):
        """Add a BoQ to a project"""
        try:
            self._init_project_collection()
            if isinstance(project_id, str):
                project_id = ObjectId(project_id)

            result = self.project_collection.update_one(
                {'_id': project_id},
                {
                    '$addToSet': {'boq_ids': boq_id},
                    '$set': {'updated_at': datetime.now(timezone.utc)}
                }
            )
            return result.modified_count > 0
        except Exception as e:
            raise Exception(f"Failed to add BoQ to project: {e}")

    def remove_boq_from_project(self, project_id, boq_id):
        """Remove a BoQ from a project"""
        try:
            self._init_project_collection()
            if isinstance(project_id, str):
                project_id = ObjectId(project_id)

            result = self.project_collection.update_one(
                {'_id': project_id},
                {
                    '$pull': {'boq_ids': boq_id},
                    '$set': {'updated_at': datetime.now(timezone.utc)}
                }
            )
            return result.modified_count > 0
        except Exception as e:
            raise Exception(f"Failed to remove BoQ from project: {e}")


class DatabaseConfigDialog(QDialog):
    """Dialog for configuring database connection"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.config_file = os.path.join(os.path.dirname(__file__), 'db_config.json')
        self.init_ui()
        self.load_saved_config()

    def init_ui(self):
        self.setWindowTitle("MongoDB Əlaqə Parametrləri")
        self.setMinimumWidth(450)

        layout = QFormLayout()

        # Host
        self.host_input = QLineEdit("")
        layout.addRow("Host:", self.host_input)

        # Port
        self.port_input = QSpinBox()
        self.port_input.setRange(1, 65535)
        self.port_input.setValue(27017)
        layout.addRow("Port:", self.port_input)

        # Database
        self.database_input = QLineEdit("admin")
        layout.addRow("Database:", self.database_input)

        # User (optional)
        self.user_input = QLineEdit("admin")
        layout.addRow("İstifadəçi:", self.user_input)

        # Password (optional)
        self.password_input = QLineEdit("")
        self.password_input.setPlaceholderText("Şifrə (istəyə bağlı)")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("Şifrə:", self.password_input)

        # Save credentials checkbox
        self.save_credentials_checkbox = QPushButton("💾 Məlumatları Yadda Saxla")
        self.save_credentials_checkbox.setCheckable(True)
        self.save_credentials_checkbox.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:checked {
                background-color: #4CAF50;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        layout.addRow(self.save_credentials_checkbox)

        # Buttons
        button_layout = QHBoxLayout()

        self.test_btn = QPushButton("🔌 Əlaqəni Yoxla")
        self.test_btn.clicked.connect(self.test_connection)
        self.test_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e68900;
            }
        """)

        self.connect_btn = QPushButton("✅ Qoşul")
        self.connect_btn.clicked.connect(self.accept)
        self.connect_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        self.cancel_btn = QPushButton("❌ Ləğv Et")
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)

        button_layout.addWidget(self.test_btn)
        button_layout.addWidget(self.connect_btn)
        button_layout.addWidget(self.cancel_btn)

        layout.addRow(button_layout)

        self.setLayout(layout)

    def test_connection(self):
        """Test the database connection with current parameters"""
        try:
            db = DatabaseManager(
                host=self.host_input.text().strip(),
                port=self.port_input.value(),
                database=self.database_input.text().strip(),
                username=self.user_input.text().strip(),
                password=self.password_input.text()
            )
            success, message = db.test_connection()
            if success:
                QMessageBox.information(
                    self,
                    "Uğurlu",
                    f"Verilənlər bazasına əlaqə uğurlu!\n\n{message}"
                )
            else:
                QMessageBox.warning(self, "Xəta", f"Əlaqə uğursuz:\n{message}")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Əlaqə xətası:\n{str(e)}")

    def load_saved_config(self):
        """Load saved configuration from file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)

                # Load values into inputs
                self.host_input.setText(config.get('host', ''))
                self.port_input.setValue(config.get('port', 27017))
                self.database_input.setText(config.get('database', 'admin'))
                self.user_input.setText(config.get('username', 'admin'))
                self.password_input.setText(config.get('password', ''))

                # Check the save button if config exists
                self.save_credentials_checkbox.setChecked(True)
        except Exception as e:
            print(f"Could not load saved config: {e}")

    def save_config(self):
        """Save configuration to file"""
        try:
            config = {
                'host': self.host_input.text().strip(),
                'port': self.port_input.value(),
                'database': self.database_input.text().strip(),
                'username': self.user_input.text().strip(),
                'password': self.password_input.text()
            }

            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Could not save config: {e}")

    def delete_saved_config(self):
        """Delete saved configuration file"""
        try:
            if os.path.exists(self.config_file):
                os.remove(self.config_file)
        except Exception as e:
            print(f"Could not delete config: {e}")

    def get_config(self):
        """Return the database configuration"""
        # Save or delete config based on checkbox state
        if self.save_credentials_checkbox.isChecked():
            self.save_config()
        else:
            self.delete_saved_config()

        return {
            'host': self.host_input.text().strip(),
            'port': self.port_input.value(),
            'database': self.database_input.text().strip(),
            'username': self.user_input.text().strip(),
            'password': self.password_input.text()
        }


class ImageViewerDialog(QDialog):
    """Dialog for viewing product images"""

    def __init__(self, parent=None, image_data=None, product_name=""):
        super().__init__(parent)
        self.image_data = image_data
        self.product_name = product_name
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"Şəkil: {self.product_name}")
        self.setMinimumSize(600, 600)

        layout = QVBoxLayout()

        # Image label
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setStyleSheet("background-color: #f0f0f0; border: 2px solid #ccc;")

        if self.image_data:
            pixmap = QPixmap()
            pixmap.loadFromData(self.image_data)

            # Scale image to fit window while maintaining aspect ratio
            scaled_pixmap = pixmap.scaled(
                800, 800,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )
            self.image_label.setPixmap(scaled_pixmap)
        else:
            self.image_label.setText("Şəkil yoxdur")
            self.image_label.setStyleSheet("color: #999; font-size: 16px;")

        # Scroll area for large images
        scroll_area = QScrollArea()
        scroll_area.setWidget(self.image_label)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Close button
        close_btn = QPushButton("❌ Bağla")
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        layout.addWidget(close_btn)

        self.setLayout(layout)


class PriceHistoryDialog(QDialog):
    """Dialog to show price change history for a product"""

    def __init__(self, parent=None, product_name="", price_history=None, current_price=0):
        super().__init__(parent)
        self.product_name = product_name
        self.price_history = price_history or []
        self.current_price = current_price
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"📈 Qiymət Tarixi - {self.product_name}")
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)

        layout = QVBoxLayout()

        # Title
        title = QLabel(f"Qiymət Tarixi: {self.product_name}")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #2196F3; padding: 10px;")
        layout.addWidget(title)

        # Current price label
        current_label = QLabel(f"Cari Qiymət: {self.current_price:.2f} AZN")
        current_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #4CAF50; padding: 5px;")
        layout.addWidget(current_label)

        # Table for history
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Tarix", "Köhnə Qiymət (AZN)", "Yeni Qiymət (AZN)", "Dəyişiklik"])
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Resize columns
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)

        # Populate table (newest first)
        sorted_history = sorted(self.price_history, key=lambda x: x.get('changed_at', datetime.min), reverse=True)

        for entry in sorted_history:
            row = self.table.rowCount()
            self.table.insertRow(row)

            # Date
            changed_at = entry.get('changed_at')
            if changed_at:
                if hasattr(changed_at, 'astimezone'):
                    local_time = changed_at.astimezone()
                    date_str = local_time.strftime("%d.%m.%Y %H:%M")
                else:
                    date_str = str(changed_at)
            else:
                date_str = "N/A"
            self.table.setItem(row, 0, QTableWidgetItem(date_str))

            # Old price
            old_price = entry.get('old_price', 0)
            self.table.setItem(row, 1, QTableWidgetItem(f"{old_price:.2f}"))

            # New price
            new_price = entry.get('new_price', 0)
            self.table.setItem(row, 2, QTableWidgetItem(f"{new_price:.2f}"))

            # Change (difference)
            diff = new_price - old_price
            diff_text = f"+{diff:.2f}" if diff >= 0 else f"{diff:.2f}"
            diff_item = QTableWidgetItem(diff_text)
            if diff > 0:
                diff_item.setForeground(QColor("#f44336"))  # Red for increase
            elif diff < 0:
                diff_item.setForeground(QColor("#4CAF50"))  # Green for decrease
            self.table.setItem(row, 3, diff_item)

        layout.addWidget(self.table)

        # Info label if no history
        if not self.price_history:
            no_history_label = QLabel("Bu məhsul üçün qiymət tarixi yoxdur.")
            no_history_label.setStyleSheet("color: #666; font-style: italic; padding: 20px;")
            no_history_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(no_history_label)

        # Close button
        close_btn = QPushButton("❌ Bağla")
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        layout.addWidget(close_btn)

        self.setLayout(layout)


class ProductDialog(QDialog):
    """Dialog for adding/editing products"""

    def __init__(self, parent=None, product=None, mode="add"):
        super().__init__(parent)
        self.product = product
        self.mode = mode
        self.parent_window = parent
        self.image_data = None  # Store uploaded image data
        self.image_filename = None
        self.remove_image = False  # Flag to indicate image removal

        # Timer for clearing status messages in dialog
        if self.mode == "add":
            self.status_clear_timer = QTimer()
            self.status_clear_timer.setSingleShot(True)
            self.status_clear_timer.timeout.connect(self._clear_dialog_status)

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Yeni Məhsul Əlavə Et" if self.mode == "add" else "Məhsulu Redaktə Et")
        self.setMinimumWidth(500)

        layout = QFormLayout()

        # Product Name
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Məhsulun adını daxil edin")
        layout.addRow("Məhsulun Adı *:", self.name_input)

        # Category
        self.category_input = QLineEdit()
        self.category_input.setPlaceholderText("Məsələn: İnşaat, Elektrik, Santexnika")
        layout.addRow("Kateqoriya:", self.category_input)

        # Price
        self.price_input = QDoubleSpinBox()
        self.price_input.setRange(0, 999999.99)
        self.price_input.setDecimals(2)
        self.price_input.setSuffix(" AZN")
        layout.addRow("Qiymət:", self.price_input)

        # Source/Origin
        self.source_input = QLineEdit()
        self.source_input.setPlaceholderText("Məhsulun mənbəyi")
        layout.addRow("Məhsul Mənbəyi:", self.source_input)

        # Unit of measurement
        self.unit_input = QLineEdit()
        self.unit_input.setPlaceholderText("kg, litr, ədəd və s.")
        layout.addRow("Ölçü Vahidi:", self.unit_input)

        # Note
        self.note_input = QTextEdit()
        self.note_input.setPlaceholderText("Əlavə qeydlər")
        self.note_input.setMaximumHeight(80)
        layout.addRow("Qeyd:", self.note_input)

        # Image upload section
        image_layout = QHBoxLayout()
        self.upload_image_btn = QPushButton("📷 Şəkil Yüklə")
        self.upload_image_btn.clicked.connect(self.upload_image)
        self.upload_image_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 6px 12px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)
        image_layout.addWidget(self.upload_image_btn)

        self.image_status_label = QLabel("Şəkil yoxdur")
        self.image_status_label.setStyleSheet("color: #666; font-style: italic;")
        image_layout.addWidget(self.image_status_label)

        self.remove_image_btn = QPushButton("🗑️ Şəkli Sil")
        self.remove_image_btn.clicked.connect(self.remove_image_action)
        self.remove_image_btn.setVisible(False)
        self.remove_image_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 6px 12px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        image_layout.addWidget(self.remove_image_btn)

        image_layout.addStretch()
        layout.addRow("Şəkil:", image_layout)

        # Fill fields if editing
        if self.product:
            self.name_input.setText(self.product['mehsulun_adi'])
            self.category_input.setText(self.product.get('category', '') or '')
            self.price_input.setValue(float(self.product['price']) if self.product.get('price') else 0)
            self.source_input.setText(self.product.get('mehsul_menbeyi', '') or '')
            self.unit_input.setText(self.product.get('olcu_vahidi', '') or '')
            self.note_input.setText(self.product.get('qeyd', '') or '')

            # Check if product has an image
            if self.product.get('image_id'):
                self.image_status_label.setText("✓ Şəkil mövcuddur")
                self.image_status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
                self.remove_image_btn.setVisible(True)

        # Status label (for add mode)
        if self.mode == "add":
            self.status_label = QLabel("")
            self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold; padding: 5px;")
            self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addRow(self.status_label)

        # Buttons
        button_layout = QHBoxLayout()

        if self.mode == "add":
            self.save_btn = QPushButton("💾 Əlavə Et və Davam Et")
            self.save_btn.clicked.connect(self.save_and_continue)
        else:
            self.save_btn = QPushButton("💾 Yadda Saxla")
            self.save_btn.clicked.connect(self.accept)

        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        self.cancel_btn = QPushButton("❌ Bağla" if self.mode == "add" else "❌ Ləğv Et")
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)

        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)

        layout.addRow(button_layout)

        self.setLayout(layout)

    def upload_image(self):
        """Handle image upload"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Şəkil Seçin",
            "",
            "Images (*.png *.jpg *.jpeg *.bmp *.gif)"
        )

        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    self.image_data = f.read()
                    self.image_filename = os.path.basename(file_path)

                self.image_status_label.setText(f"✓ {self.image_filename}")
                self.image_status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
                self.remove_image_btn.setVisible(True)
                self.remove_image = False
            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Şəkil yüklənə bilmədi: {e}")

    def remove_image_action(self):
        """Mark image for removal"""
        self.image_data = None
        self.image_filename = None
        self.remove_image = True
        self.image_status_label.setText("Şəkil yoxdur")
        self.image_status_label.setStyleSheet("color: #666; font-style: italic;")
        self.remove_image_btn.setVisible(False)

    def get_data(self):
        """Returns the form data"""
        data = {
            'mehsulun_adi': self.name_input.text().strip(),
            'category': self.category_input.text().strip() or None,
            'price': self.price_input.value() if self.price_input.value() > 0 else None,
            'mehsul_menbeyi': self.source_input.text().strip() or None,
            'qeyd': self.note_input.toPlainText().strip() or None,
            'olcu_vahidi': self.unit_input.text().strip() or None
        }

        # Add image information
        data['image_data'] = self.image_data
        data['image_filename'] = self.image_filename
        data['remove_image'] = self.remove_image

        return data

    def save_and_continue(self):
        """Save product and clear form for adding another (add mode only)"""
        if not self.name_input.text().strip():
            self.status_label.setText("⚠️ Məhsulun adı mütləqdir!")
            self.status_label.setStyleSheet("color: #f44336; font-weight: bold; padding: 5px;")
            return

        data = self.get_data()
        try:
            # Call parent window's database to create product
            if self.parent_window and self.parent_window.db:
                # Handle image upload
                image_id = None
                if data['image_data']:
                    image_id = self.parent_window.db.save_image(data['image_data'], data['image_filename'])

                product_id = self.parent_window.db.create_product(
                    data['mehsulun_adi'],
                    data['price'],
                    data['mehsul_menbeyi'],
                    data['qeyd'],
                    data['olcu_vahidi'],
                    data['category'],
                    image_id=image_id
                )

                # Show success status in dialog
                self.status_label.setText(f"✅ '{data['mehsulun_adi']}' uğurla əlavə edildi!")
                self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold; padding: 5px;")

                # Auto-clear dialog status after 2 seconds
                self.status_clear_timer.stop()
                self.status_clear_timer.start(2000)

                # Update parent window's table
                self.parent_window.load_products(preserve_status=True)

                # Show status in main window
                self.parent_window.show_status(
                    f"✅ '{data['mehsulun_adi']}' məhsulu əlavə edildi",
                    color="#4CAF50"
                )

                # Clear form for next entry
                self.clear_form()

                # Set focus back to name field
                self.name_input.setFocus()

        except Exception as e:
            self.status_label.setText(f"❌ Xəta: {str(e)}")
            self.status_label.setStyleSheet("color: #f44336; font-weight: bold; padding: 5px;")
            # Auto-clear error after 4 seconds
            self.status_clear_timer.stop()
            self.status_clear_timer.start(4000)

    def clear_form(self):
        """Clear all input fields"""
        self.name_input.clear()
        self.category_input.clear()
        self.price_input.setValue(0)
        self.source_input.clear()
        self.unit_input.clear()
        self.note_input.clear()

        # Clear image data
        self.image_data = None
        self.image_filename = None
        self.remove_image = False
        self.image_status_label.setText("Şəkil yoxdur")
        self.image_status_label.setStyleSheet("color: #666; font-style: italic;")
        self.remove_image_btn.setVisible(False)

    def _clear_dialog_status(self):
        """Clear the status label in the dialog"""
        if hasattr(self, 'status_label'):
            self.status_label.setText("")

    def accept(self):
        """Validate before accepting (edit mode)"""
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Xəbərdarlıq", "Məhsulun adı mütləqdir!")
            return
        super().accept()


class BoQItemDialog(QDialog):
    """Dialog for adding items to BoQ (from DB or custom)"""

    def __init__(self, parent=None, db=None, item=None, mode="add_from_db"):
        super().__init__(parent)
        self.db = db
        self.item = item
        self.mode = mode  # "add_from_db", "custom", "edit"
        self.init_ui()

    def init_ui(self):
        if self.mode == "add_from_db":
            self.setWindowTitle("Məhsul Əlavə Et (Verilənlər Bazasından)")
        elif self.mode == "custom":
            self.setWindowTitle("Xüsusi Qeyd Əlavə Et")
        else:
            self.setWindowTitle("Qeydi Redaktə Et")

        self.setMinimumWidth(500)
        layout = QFormLayout()

        # Product selection (only for add_from_db mode)
        if self.mode == "add_from_db":
            self.product_combo = QLineEdit()
            self.product_combo.setPlaceholderText("Məhsul ID-ni daxil edin")
            layout.addRow("Məhsul ID:", self.product_combo)

        # Name (editable for custom, read-only for DB items)
        self.name_input = QLineEdit()
        if self.mode == "add_from_db":
            self.name_input.setReadOnly(True)
            self.name_input.setPlaceholderText("ID daxil etdikdən sonra avtomatik doldurulacaq")
        else:
            self.name_input.setPlaceholderText("Məhsul/İş adını daxil edin")
        layout.addRow("Adı:", self.name_input)

        # Quantity
        self.quantity_input = QDoubleSpinBox()
        self.quantity_input.setRange(0.01, 999999.99)
        self.quantity_input.setDecimals(2)
        self.quantity_input.setValue(1)
        layout.addRow("Miqdar:", self.quantity_input)

        # Unit
        self.unit_input = QLineEdit()
        if self.mode == "add_from_db":
            self.unit_input.setReadOnly(True)
        else:
            self.unit_input.setPlaceholderText("kg, m², ədəd və s.")
        layout.addRow("Ölçü Vahidi:", self.unit_input)

        # Unit Price
        self.price_input = QDoubleSpinBox()
        self.price_input.setRange(0, 999999.99)
        self.price_input.setDecimals(2)
        self.price_input.setSuffix(" AZN")
        if self.mode == "add_from_db":
            self.price_input.setReadOnly(True)
        layout.addRow("Vahid Qiymət:", self.price_input)

        # Total (calculated, read-only)
        self.total_label = QLabel("0.00 AZN")
        self.total_label.setStyleSheet("font-weight: bold; color: #2196F3;")
        layout.addRow("Cəmi:", self.total_label)

        # Margin percent
        self.margin_input = QDoubleSpinBox()
        self.margin_input.setRange(0, 100)
        self.margin_input.setDecimals(1)
        self.margin_input.setValue(0)
        self.margin_input.setSuffix(" %")
        layout.addRow("Marja %:", self.margin_input)

        # Final price with margin (calculated, read-only)
        self.final_total_label = QLabel("0.00 AZN")
        self.final_total_label.setStyleSheet("font-weight: bold; color: #4CAF50;")
        layout.addRow("Yekun (Marja ilə):", self.final_total_label)

        # Connect quantity change to update total
        self.quantity_input.valueChanged.connect(self.update_total)
        self.price_input.valueChanged.connect(self.update_total)
        self.margin_input.valueChanged.connect(self.update_total)

        # Load product info button (for add_from_db mode)
        if self.mode == "add_from_db":
            self.load_btn = QPushButton("📥 Məhsul Məlumatını Yüklə")
            self.load_btn.clicked.connect(self.load_product_info)
            self.load_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #0b7dda;
                }
            """)
            layout.addRow(self.load_btn)

        # Fill fields if editing
        if self.item:
            self.name_input.setText(self.item.get('name', ''))
            self.quantity_input.setValue(float(self.item.get('quantity', 1)))
            self.unit_input.setText(self.item.get('unit', ''))
            self.price_input.setValue(float(self.item.get('unit_price', 0)))
            self.margin_input.setValue(float(self.item.get('margin_percent', 0)))
            self.update_total()

        # Buttons
        button_layout = QHBoxLayout()

        self.save_btn = QPushButton("💾 Əlavə Et" if self.mode != "edit" else "💾 Yadda Saxla")
        self.save_btn.clicked.connect(self.accept)
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        self.cancel_btn = QPushButton("❌ Ləğv Et")
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)

        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addRow(button_layout)

        self.setLayout(layout)

    def load_product_info(self):
        """Load product information from database"""
        try:
            product_id = self.product_combo.text().strip()
            product = self.db.read_product(product_id)

            if product:
                self.name_input.setText(product['mehsulun_adi'])
                self.unit_input.setText(product.get('olcu_vahidi', '') or 'ədəd')
                self.price_input.setValue(float(product['price']) if product.get('price') else 0)
                self.update_total()
                QMessageBox.information(self, "Uğurlu", "Məhsul məlumatı yükləndi!")
            else:
                QMessageBox.warning(self, "Xəta", "Bu ID-yə sahib məhsul tapılmadı!")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Xəta baş verdi: {str(e)}")

    def update_total(self):
        """Update total price and final total with margin"""
        quantity = self.quantity_input.value()
        unit_price = self.price_input.value()
        total = quantity * unit_price
        margin_percent = self.margin_input.value()
        final_total = total * (1 + margin_percent / 100)
        self.total_label.setText(f"{total:.2f} AZN")
        self.final_total_label.setText(f"{final_total:.2f} AZN")

    def get_data(self):
        """Get form data"""
        return {
            'product_id': self.product_combo.text().strip() if self.mode == "add_from_db" and hasattr(self, 'product_combo') and self.product_combo.text().strip() else None,
            'name': self.name_input.text().strip(),
            'quantity': self.quantity_input.value(),
            'unit': self.unit_input.text().strip(),
            'unit_price': self.price_input.value(),
            'total': self.quantity_input.value() * self.price_input.value(),
            'margin_percent': self.margin_input.value(),
            'is_custom': self.mode == "custom"
        }

    def accept(self):
        """Validate before accepting"""
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Xəbərdarlıq", "Ad mütləqdir!")
            return
        if self.quantity_input.value() <= 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Miqdar 0-dan böyük olmalıdır!")
            return
        super().accept()


class BoQWindow(QMainWindow):
    """Bill of Quantities Window"""

    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.parent_window = parent
        self.db = db
        self.boq_items = []  # List to store BoQ items
        self.next_id = 1
        self.boq_name = "BoQ 1"  # Default name
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("📋 Bill of Quantities (BoQ)")
        self.setGeometry(150, 150, 1200, 650)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()

        # Title and Name section with Up/Down buttons
        title_layout = QHBoxLayout()

        title = QLabel("Bill of Quantities")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #2196F3; padding: 10px;")

        # Move Up/Down buttons (small, at top)
        self.move_up_btn = QPushButton("⬆️")
        self.move_up_btn.clicked.connect(self.move_item_up)
        self.move_up_btn.setFixedSize(40, 30)
        self.move_up_btn.setStyleSheet("background-color: #607D8B; color: white; border: none; border-radius: 4px; font-weight: bold;")
        self.move_up_btn.setToolTip("Yuxarı daşı (Ctrl+Up)")

        self.move_down_btn = QPushButton("⬇️")
        self.move_down_btn.clicked.connect(self.move_item_down)
        self.move_down_btn.setFixedSize(40, 30)
        self.move_down_btn.setStyleSheet("background-color: #607D8B; color: white; border: none; border-radius: 4px; font-weight: bold;")
        self.move_down_btn.setToolTip("Aşağı daşı (Ctrl+Down)")

        # BoQ Name input
        name_label = QLabel("BoQ Adı:")
        name_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        self.boq_name_input = QLineEdit(self.boq_name)
        self.boq_name_input.setMaximumWidth(200)
        self.boq_name_input.setStyleSheet("font-size: 14px; padding: 5px;")
        self.boq_name_input.textChanged.connect(self.update_boq_name)

        title_layout.addWidget(title)
        title_layout.addWidget(self.move_up_btn)
        title_layout.addWidget(self.move_down_btn)
        title_layout.addStretch()
        title_layout.addWidget(name_label)
        title_layout.addWidget(self.boq_name_input)

        main_layout.addLayout(title_layout)

        # Table with margin column
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "№", "Adı", "Kateqoriya", "Miqdar", "Ölçü Vahidi", "Vahid Qiymət", "Cəmi", "Marja %", "Yekun", "Mənbə", "Qeyd", "Növ"
        ])

        # Table styling
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)  # Enable column sorting

        # Resize columns
        header = self.table.horizontalHeader()
        header.setSortIndicatorShown(True)
        header.sectionClicked.connect(self.sort_by_column)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # №
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Adı
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Kateqoriya
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Miqdar
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Ölçü Vahidi
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Vahid Qiymət
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Cəmi
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)  # Marja %
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)  # Yekun
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)  # Mənbə
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)  # Qeyd
        header.setSectionResizeMode(11, QHeaderView.ResizeMode.ResizeToContents)  # Növ

        main_layout.addWidget(self.table)

        # Summary labels
        summary_layout = QHBoxLayout()
        summary_layout.addStretch()

        self.cost_label = QLabel("Maya Dəyəri: 0.00 AZN")
        self.cost_label.setStyleSheet("font-size: 14px; color: #666; padding: 10px;")

        self.margin_total_label = QLabel("Ümumi Marja: 0.00 AZN")
        self.margin_total_label.setStyleSheet("font-size: 14px; color: #FF9800; padding: 10px;")

        self.summary_label = QLabel("Yekun Məbləğ: 0.00 AZN")
        self.summary_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #4CAF50; padding: 10px;")

        summary_layout.addWidget(self.cost_label)
        summary_layout.addWidget(self.margin_total_label)
        summary_layout.addWidget(self.summary_label)

        main_layout.addLayout(summary_layout)

        # Buttons - reorganized into two rows
        button_layout1 = QHBoxLayout()
        button_layout2 = QHBoxLayout()

        self.add_custom_btn = QPushButton("➕ Xüsusi Qeyd")
        self.add_custom_btn.clicked.connect(self.add_custom_item)
        self.add_custom_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px 15px;
                border: none;
                border-radius: 5px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)

        self.edit_btn = QPushButton("✏️ Redaktə Et")
        self.edit_btn.clicked.connect(self.edit_item)
        self.edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e68900;
            }
        """)

        self.delete_btn = QPushButton("🗑️ Sil")
        self.delete_btn.clicked.connect(self.delete_item)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)

        self.save_boq_btn = QPushButton("💾 BoQ Yadda Saxla")
        self.save_boq_btn.clicked.connect(self.save_boq)
        self.save_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)

        self.load_boq_btn = QPushButton("📂 BoQ Yüklə")
        self.load_boq_btn.clicked.connect(self.load_boq)
        self.load_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e68900;
            }
        """)

        self.export_excel_btn = QPushButton("📊 Excel-ə İxrac Et")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        self.load_cloud_boq_btn = QPushButton("☁️ Buluddan Yüklə")
        self.load_cloud_boq_btn.clicked.connect(self.load_from_cloud)
        self.load_cloud_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #00BCD4;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0097A7;
            }
        """)

        self.combine_boq_btn = QPushButton("🔗 BoQ-ları Birləşdir")
        self.combine_boq_btn.clicked.connect(self.combine_boqs_to_excel)
        self.combine_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #673AB7;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5E35B1;
            }
        """)

        # Template Management button
        self.template_mgmt_btn = QPushButton("📋 Şablonlar")
        self.template_mgmt_btn.clicked.connect(self.open_template_management)
        self.template_mgmt_btn.setStyleSheet("""
            QPushButton {
                background-color: #795548;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #6D4C41;
            }
        """)

        # Row 1: Edit actions
        button_layout1.addWidget(self.add_custom_btn)
        button_layout1.addWidget(self.edit_btn)
        button_layout1.addWidget(self.delete_btn)
        button_layout1.addWidget(self.save_boq_btn)
        button_layout1.addWidget(self.load_boq_btn)
        button_layout1.addStretch()

        # Row 2: Export/Cloud/Templates
        button_layout2.addWidget(self.load_cloud_boq_btn)
        button_layout2.addWidget(self.export_excel_btn)
        button_layout2.addWidget(self.combine_boq_btn)
        button_layout2.addWidget(self.template_mgmt_btn)
        button_layout2.addStretch()

        main_layout.addLayout(button_layout1)
        main_layout.addLayout(button_layout2)
        central_widget.setLayout(main_layout)

        # Apply light mode styling
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
        """)

        # Setup keyboard shortcuts
        self.setup_shortcuts()

    def setup_shortcuts(self):
        """Setup keyboard shortcuts for BoQWindow"""
        # Ctrl+S: Save BoQ
        QShortcut(QKeySequence("Ctrl+S"), self).activated.connect(self.save_boq)

        # Ctrl+O: Load BoQ
        QShortcut(QKeySequence("Ctrl+O"), self).activated.connect(self.load_boq)

        # Ctrl+E: Edit selected item
        QShortcut(QKeySequence("Ctrl+E"), self).activated.connect(self.edit_item)

        # Delete: Delete selected item(s)
        QShortcut(QKeySequence("Delete"), self).activated.connect(self.delete_item)

        # Ctrl+Up: Move item up
        QShortcut(QKeySequence("Ctrl+Up"), self).activated.connect(self.move_item_up)

        # Ctrl+Down: Move item down
        QShortcut(QKeySequence("Ctrl+Down"), self).activated.connect(self.move_item_down)

    def add_from_database(self):
        """Add item from database"""
        if not self.db:
            QMessageBox.warning(self, "Xəta", "Verilənlər bazasına qoşulmayıbsınız!")
            return

        dialog = BoQItemDialog(self, self.db, mode="add_from_db")
        if dialog.exec():
            data = dialog.get_data()
            data['id'] = self.next_id
            self.next_id += 1
            self.boq_items.append(data)
            self.refresh_table()

    def add_custom_item(self):
        """Add custom item (not from database)"""
        dialog = BoQItemDialog(self, self.db, mode="custom")
        if dialog.exec():
            data = dialog.get_data()
            data['id'] = self.next_id
            self.next_id += 1
            self.boq_items.append(data)
            self.refresh_table()

    def edit_item(self):
        """Edit selected item"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Redaktə etmək üçün qeyd seçin!")
            return

        item = self.boq_items[selected_row]
        mode = "custom" if item.get('is_custom') else "edit"

        dialog = BoQItemDialog(self, self.db, item=item, mode=mode)
        if dialog.exec():
            data = dialog.get_data()
            data['id'] = item['id']
            self.boq_items[selected_row] = data
            self.refresh_table()

    def delete_item(self):
        """Delete selected item(s)"""
        # Get all selected rows
        selected_rows = self.table.selectionModel().selectedRows()

        if not selected_rows:
            QMessageBox.warning(self, "Xəbərdarlıq", "Silmək üçün qeyd seçin!")
            return

        # Confirm deletion
        if len(selected_rows) == 1:
            message = "Bu qeydi silmək istədiyinizdən əminsiniz?"
        else:
            message = f"{len(selected_rows)} qeydi silmək istədiyinizdən əminsiniz?"

        reply = QMessageBox.question(
            self,
            "Təsdiq",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Get row indices and sort in descending order
            # This allows us to delete from bottom to top without affecting other indices
            rows_to_delete = sorted([index.row() for index in selected_rows], reverse=True)

            # Delete items from the list
            for row in rows_to_delete:
                del self.boq_items[row]

            self.refresh_table()

    def move_item_up(self):
        """Move selected item up in the list"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Yuxarı daşımaq üçün qeyd seçin!")
            return

        if selected_row == 0:
            return  # Already at top

        # Swap items
        self.boq_items[selected_row], self.boq_items[selected_row - 1] = \
            self.boq_items[selected_row - 1], self.boq_items[selected_row]

        # Renumber items
        self._renumber_items()
        self.refresh_table()

        # Keep selection on the moved item
        self.table.selectRow(selected_row - 1)

    def move_item_down(self):
        """Move selected item down in the list"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Aşağı daşımaq üçün qeyd seçin!")
            return

        if selected_row >= len(self.boq_items) - 1:
            return  # Already at bottom

        # Swap items
        self.boq_items[selected_row], self.boq_items[selected_row + 1] = \
            self.boq_items[selected_row + 1], self.boq_items[selected_row]

        # Renumber items
        self._renumber_items()
        self.refresh_table()

        # Keep selection on the moved item
        self.table.selectRow(selected_row + 1)

    def _renumber_items(self):
        """Renumber all items sequentially"""
        for idx, item in enumerate(self.boq_items, start=1):
            item['id'] = idx
        self.next_id = len(self.boq_items) + 1

    def refresh_table(self):
        """Refresh the BoQ table"""
        self.table.setRowCount(0)

        for item in self.boq_items:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            # Get margin percent (default 0)
            margin_pct = item.get('margin_percent', 0)
            cost_total = item['total']
            final_total = cost_total * (1 + margin_pct / 100)

            # Column 0: №
            self.table.setItem(row_position, 0, QTableWidgetItem(str(item['id'])))
            # Column 1: Adı
            self.table.setItem(row_position, 1, QTableWidgetItem(item['name']))
            # Column 2: Kateqoriya
            self.table.setItem(row_position, 2, QTableWidgetItem(item.get('category', '') or 'N/A'))
            # Column 3: Miqdar
            self.table.setItem(row_position, 3, QTableWidgetItem(f"{item['quantity']:.2f}"))
            # Column 4: Ölçü Vahidi
            self.table.setItem(row_position, 4, QTableWidgetItem(item['unit']))
            # Column 5: Vahid Qiymət
            self.table.setItem(row_position, 5, QTableWidgetItem(f"{item['unit_price']:.2f}"))
            # Column 6: Cəmi (cost)
            self.table.setItem(row_position, 6, QTableWidgetItem(f"{cost_total:.2f}"))
            # Column 7: Marja %
            self.table.setItem(row_position, 7, QTableWidgetItem(f"{margin_pct:.1f}%"))
            # Column 8: Yekun (with margin)
            self.table.setItem(row_position, 8, QTableWidgetItem(f"{final_total:.2f}"))
            # Column 9: Mənbə
            self.table.setItem(row_position, 9, QTableWidgetItem(item.get('source', '') or 'N/A'))
            # Column 10: Qeyd
            self.table.setItem(row_position, 10, QTableWidgetItem(item.get('note', '') or 'N/A'))
            # Column 11: Növ
            item_type = "Xüsusi" if item.get('is_custom') else "DB"
            self.table.setItem(row_position, 11, QTableWidgetItem(item_type))

        # Update summary
        self.update_summary()

    def update_summary(self):
        """Update the summary labels with cost total, margin total, and final amount"""
        cost_total = sum(item['total'] for item in self.boq_items)
        margin_total = sum(item['total'] * (item.get('margin_percent', 0) / 100) for item in self.boq_items)
        final_total = cost_total + margin_total

        self.cost_label.setText(f"Maya Dəyəri: {cost_total:.2f} AZN")
        self.margin_total_label.setText(f"Ümumi Marja: {margin_total:.2f} AZN")
        self.summary_label.setText(f"Yekun Məbləğ: {final_total:.2f} AZN")

    def sort_by_column(self, column):
        """Sort boq_items by the clicked column"""
        # Map column indices to item keys
        column_keys = {
            0: 'id',
            1: 'name',
            2: 'category',
            3: 'quantity',
            4: 'unit',
            5: 'unit_price',
            6: 'total',
            7: 'margin_percent',
            8: 'final_total',  # computed, not stored
            9: 'source',
            10: 'note',
            11: 'is_custom'
        }

        key = column_keys.get(column)
        if not key:
            return

        # Check current sort order from header
        header = self.table.horizontalHeader()
        current_order = header.sortIndicatorOrder()

        # Sort the boq_items list
        reverse = (current_order == Qt.SortOrder.DescendingOrder)

        try:
            if key == 'final_total':
                # Computed field: total * (1 + margin_percent/100)
                self.boq_items.sort(key=lambda x: x.get('total', 0) * (1 + x.get('margin_percent', 0) / 100), reverse=reverse)
            elif key in ['quantity', 'unit_price', 'total', 'id', 'margin_percent']:
                # Numeric sort
                self.boq_items.sort(key=lambda x: float(x.get(key, 0) or 0), reverse=reverse)
            elif key == 'is_custom':
                # Boolean sort
                self.boq_items.sort(key=lambda x: x.get(key, False), reverse=reverse)
            else:
                # String sort
                self.boq_items.sort(key=lambda x: str(x.get(key, '') or '').lower(), reverse=reverse)

            # Renumber after sort
            self._renumber_items()

            # Refresh table without triggering sort again
            self.table.setSortingEnabled(False)
            self.refresh_table()
            self.table.setSortingEnabled(True)

        except Exception as e:
            print(f"Sort error: {e}")

    def export_to_excel(self):
        """Export BoQ to Excel file"""
        if not self.boq_items:
            QMessageBox.warning(self, "Xəbərdarlıq", "BoQ boşdur! İxrac etmək üçün məhsul əlavə edin.")
            return

        try:
            # Import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                QMessageBox.critical(
                    self,
                    "Xəta",
                    "openpyxl kitabxanası tapılmadı!\n\nYükləmək üçün terminal-da:\npip install openpyxl"
                )
                return

            # Ask user for file location
            from PyQt6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "BoQ-u Excel-ə İxrac Et",
                "BoQ.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bill of Quantities"

            # Define styles
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF", size=12)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add title
            ws.merge_cells('A1:L1')
            ws['A1'] = "BILL OF QUANTITIES (BOQ)"
            ws['A1'].font = Font(bold=True, size=16, color="2196F3")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Margin header style
            margin_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")

            # Add headers
            headers = ["№", "Adı", "Kateqoriya", "Miqdar", "Ölçü Vahidi", "Vahid Qiymət (AZN)", "Cəmi (AZN)", "Marja %", "Yekun (AZN)", "Mənbə", "Qeyd", "Növ"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                # Use orange color for margin columns
                if col_num in [8, 9]:
                    cell.fill = margin_fill
                else:
                    cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Add data
            cost_total = 0
            margin_total = 0
            final_total = 0
            for row_num, item in enumerate(self.boq_items, 4):
                margin_pct = item.get('margin_percent', 0)
                item_cost = item['total']
                item_final = item_cost * (1 + margin_pct / 100)

                # Column 1: №
                ws.cell(row=row_num, column=1, value=item['id']).border = border
                # Column 2: Adı
                ws.cell(row=row_num, column=2, value=item['name']).border = border
                # Column 3: Kateqoriya
                ws.cell(row=row_num, column=3, value=item.get('category', '') or 'N/A').border = border
                # Column 4: Miqdar
                ws.cell(row=row_num, column=4, value=item['quantity']).border = border
                # Column 5: Ölçü Vahidi
                ws.cell(row=row_num, column=5, value=item['unit']).border = border
                # Column 6: Vahid Qiymət
                ws.cell(row=row_num, column=6, value=item['unit_price']).border = border
                # Column 7: Cəmi (cost)
                ws.cell(row=row_num, column=7, value=item_cost).border = border
                # Column 8: Marja %
                ws.cell(row=row_num, column=8, value=margin_pct).border = border
                # Column 9: Yekun (with margin)
                ws.cell(row=row_num, column=9, value=item_final).border = border
                # Column 10: Mənbə
                ws.cell(row=row_num, column=10, value=item.get('source', '') or 'N/A').border = border
                # Column 11: Qeyd
                ws.cell(row=row_num, column=11, value=item.get('note', '') or 'N/A').border = border
                # Column 12: Növ
                item_type = "Xüsusi" if item.get('is_custom') else "DB"
                ws.cell(row=row_num, column=12, value=item_type).border = border

                # Format numbers
                ws.cell(row=row_num, column=4).number_format = '0.00'
                ws.cell(row=row_num, column=6).number_format = '0.00'
                ws.cell(row=row_num, column=7).number_format = '0.00'
                ws.cell(row=row_num, column=8).number_format = '0.0'
                ws.cell(row=row_num, column=9).number_format = '0.00'

                cost_total += item_cost
                margin_total += (item_final - item_cost)
                final_total += item_final

            # Add total row
            total_row = len(self.boq_items) + 5

            # Cost total label
            ws.merge_cells(f'A{total_row}:F{total_row}')
            cost_label_cell = ws[f'A{total_row}']
            cost_label_cell.value = "MAYA DƏYƏRİ:"
            cost_label_cell.fill = header_fill
            cost_label_cell.font = total_font
            cost_label_cell.alignment = Alignment(horizontal="right", vertical="center")
            cost_label_cell.border = border

            # Cost total value
            cost_value_cell = ws[f'G{total_row}']
            cost_value_cell.value = cost_total
            cost_value_cell.fill = header_fill
            cost_value_cell.font = total_font
            cost_value_cell.alignment = Alignment(horizontal="center", vertical="center")
            cost_value_cell.border = border
            cost_value_cell.number_format = '0.00'

            # Margin total label
            margin_label_cell = ws[f'H{total_row}']
            margin_label_cell.value = "Marja:"
            margin_label_cell.fill = margin_fill
            margin_label_cell.font = total_font
            margin_label_cell.alignment = Alignment(horizontal="center", vertical="center")
            margin_label_cell.border = border

            # Final total value
            final_value_cell = ws[f'I{total_row}']
            final_value_cell.value = final_total
            final_value_cell.fill = total_fill
            final_value_cell.font = total_font
            final_value_cell.alignment = Alignment(horizontal="center", vertical="center")
            final_value_cell.border = border
            final_value_cell.number_format = '0.00'

            # Fill remaining cells in total row
            for col in ['J', 'K', 'L']:
                ws[f'{col}{total_row}'].fill = total_fill
                ws[f'{col}{total_row}'].border = border

            # Add margin summary row
            margin_row = total_row + 1
            ws.merge_cells(f'A{margin_row}:F{margin_row}')
            margin_summary_cell = ws[f'A{margin_row}']
            margin_summary_cell.value = "ÜMUMİ MARJA:"
            margin_summary_cell.fill = margin_fill
            margin_summary_cell.font = total_font
            margin_summary_cell.alignment = Alignment(horizontal="right", vertical="center")
            margin_summary_cell.border = border

            margin_summary_value = ws[f'G{margin_row}']
            margin_summary_value.value = margin_total
            margin_summary_value.fill = margin_fill
            margin_summary_value.font = total_font
            margin_summary_value.alignment = Alignment(horizontal="center", vertical="center")
            margin_summary_value.border = border
            margin_summary_value.number_format = '0.00'

            # Fill remaining cells
            for col in ['H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{margin_row}'].fill = margin_fill
                ws[f'{col}{margin_row}'].border = border

            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 10

            # Save file
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Uğurlu",
                f"BoQ uğurla Excel-ə ixrac edildi!\n\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"İxrac zamanı xəta:\n{str(e)}")

    def update_boq_name(self):
        """Update BoQ name from input field"""
        self.boq_name = self.boq_name_input.text().strip() or "BoQ 1"

    def save_boq(self):
        """Save BoQ to JSON file with optional cloud save"""
        if not self.boq_items:
            QMessageBox.warning(self, "Xəbərdarlıq", "BoQ boşdur! Yadda saxlamaq üçün məhsul əlavə edin.")
            return

        try:
            import json
            from PyQt6.QtWidgets import QFileDialog, QCheckBox

            # Create custom dialog with checkbox
            dialog = QDialog(self)
            dialog.setWindowTitle("BoQ-u Yadda Saxla")
            dialog.setMinimumWidth(400)

            layout = QVBoxLayout()

            # Info label
            info_label = QLabel("BoQ-u harada saxlamaq istəyirsiniz?")
            layout.addWidget(info_label)

            # Cloud save checkbox
            cloud_checkbox = QCheckBox("☁️ Buludda da saxla (MongoDB)")
            cloud_checkbox.setChecked(True)  # Default checked
            cloud_checkbox.setStyleSheet("padding: 10px; font-size: 12px;")
            layout.addWidget(cloud_checkbox)

            # Buttons
            button_layout = QHBoxLayout()
            save_btn = QPushButton("💾 Yadda Saxla")
            save_btn.setDefault(True)
            cancel_btn = QPushButton("❌ Ləğv Et")

            save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)

            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
            """)

            button_layout.addWidget(save_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)

            dialog.setLayout(layout)

            # Connect buttons
            save_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            # Show dialog
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            save_to_cloud = cloud_checkbox.isChecked()

            # Ask user for file location
            default_name = f"{self.boq_name}.json" if self.boq_name else "BoQ.json"

            # Set default directory to saved_boqs folder
            saved_boqs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "saved_boqs")
            os.makedirs(saved_boqs_dir, exist_ok=True)
            default_path = os.path.join(saved_boqs_dir, default_name)

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "BoQ-u Yadda Saxla",
                default_path,
                "JSON Files (*.json)"
            )

            if not file_path:
                return

            # Prepare data for saving
            save_data = {
                'boq_name': self.boq_name,
                'next_id': self.next_id,
                'items': self.boq_items
            }

            # Save to local file
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)

            success_message = f"BoQ lokal faylda yadda saxlanıldı:\n{file_path}"

            # Save to cloud if checkbox is checked
            if save_to_cloud and self.db:
                try:
                    _, is_new = self.db.save_boq_to_cloud(
                        self.boq_name,
                        self.boq_items,
                        self.next_id
                    )
                    if is_new:
                        success_message += "\n\n✅ Buludda da saxlanıldı (yeni)!"
                    else:
                        success_message += "\n\n✅ Buludda yeniləndi!"
                except Exception as cloud_error:
                    success_message += f"\n\n⚠️ Bulud saxlama xətası: {cloud_error}"

            QMessageBox.information(
                self,
                "Uğurlu",
                success_message
            )

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"BoQ yadda saxlanarkən xəta:\n{str(e)}")

    def load_boq(self):
        """Load BoQ from JSON file and update prices from database"""
        try:
            import json
            from PyQt6.QtWidgets import QFileDialog

            # Ask user for file to load
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "BoQ Yüklə",
                "",
                "JSON Files (*.json)"
            )

            if not file_path:
                return

            # Confirm if current BoQ will be replaced
            if self.boq_items:
                reply = QMessageBox.question(
                    self,
                    "Təsdiq",
                    "Mövcud BoQ məlumatları əvəz olunacaq. Davam etmək istəyirsiniz?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return

            # Load from file
            with open(file_path, 'r', encoding='utf-8') as f:
                save_data = json.load(f)

            self.boq_name = save_data.get('boq_name', 'BoQ 1')
            self.boq_name_input.setText(self.boq_name)
            self.next_id = save_data.get('next_id', 1)
            loaded_items = save_data.get('items', [])

            # Update prices from database for items that came from DB
            updated_count = 0
            for item in loaded_items:
                if not item.get('is_custom') and item.get('product_id') and self.db:
                    try:
                        # Get current product data from database
                        product = self.db.read_product(item['product_id'])
                        if product:
                            # Update price, category, source, and note from database
                            old_price = item['unit_price']
                            new_price = float(product['price']) if product.get('price') else 0

                            item['unit_price'] = new_price
                            item['total'] = item['quantity'] * new_price
                            item['category'] = product.get('category', '') or ''
                            item['source'] = product.get('mehsul_menbeyi', '') or ''
                            item['note'] = product.get('qeyd', '') or ''

                            if old_price != new_price:
                                updated_count += 1
                    except Exception as e:
                        # If product not found or error, keep the saved data
                        pass

            self.boq_items = loaded_items
            self.refresh_table()

            message = f"BoQ uğurla yükləndi!\n\n{len(loaded_items)} məhsul yükləndi."
            if updated_count > 0:
                message += f"\n{updated_count} məhsulun qiyməti yeniləndi."

            QMessageBox.information(self, "Uğurlu", message)

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"BoQ yüklənərkən xəta:\n{str(e)}")

    def load_from_cloud(self):
        """Load BoQ from cloud (MongoDB)"""
        if not self.db:
            QMessageBox.warning(self, "Xəbərdarlıq", "Verilənlər bazasına qoşulun!")
            return

        try:
            # Get all cloud BoQs
            cloud_boqs = self.db.get_all_cloud_boqs()

            if not cloud_boqs:
                QMessageBox.information(self, "Məlumat", "Buludda heç bir BoQ tapılmadı!")
                return

            # Create selection dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Buluddan BoQ Yüklə")
            dialog.setMinimumWidth(600)
            dialog.setMinimumHeight(500)

            layout = QVBoxLayout()

            # Title
            title = QLabel("Yükləmək üçün BoQ seçin:")
            title.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px;")
            layout.addWidget(title)

            # Search input
            search_layout = QHBoxLayout()
            search_input = QLineEdit()
            search_input.setPlaceholderText("🔍 BoQ adına görə axtar...")
            search_input.setStyleSheet("""
                QLineEdit {
                    padding: 8px 12px;
                    border: 2px solid #ddd;
                    border-radius: 5px;
                    font-size: 13px;
                }
                QLineEdit:focus {
                    border-color: #2196F3;
                }
            """)
            search_layout.addWidget(search_input)
            layout.addLayout(search_layout)

            # List widget for BoQs
            from PyQt6.QtWidgets import QListWidget, QListWidgetItem
            boq_list = QListWidget()
            boq_list.setStyleSheet("""
                QListWidget {
                    font-size: 13px;
                    padding: 5px;
                }
                QListWidget::item {
                    padding: 10px;
                    border-bottom: 1px solid #ddd;
                }
                QListWidget::item:hover {
                    background-color: #e3f2fd;
                }
                QListWidget::item:selected {
                    background-color: #2196F3;
                    color: white;
                }
            """)

            # Store all BoQs for filtering
            all_boqs = cloud_boqs

            def format_boq_item(boq):
                """Format BoQ data for display"""
                item_count = len(boq.get('items', []))
                updated_at = boq.get('updated_at', '')
                if updated_at:
                    try:
                        # Convert UTC to local time
                        if updated_at.tzinfo is None:
                            updated_at = updated_at.replace(tzinfo=timezone.utc)
                        local_time = updated_at.astimezone()
                        updated_str = local_time.strftime('%Y-%m-%d %H:%M')
                    except:
                        updated_str = str(updated_at)
                else:
                    updated_str = "Naməlum"
                return f"{boq['name']} ({item_count} məhsul) - Son yenilənmə: {updated_str}"

            def populate_list(boqs):
                """Populate the list with BoQs"""
                boq_list.clear()
                for boq in boqs:
                    item_text = format_boq_item(boq)
                    list_item = QListWidgetItem(item_text)
                    list_item.setData(Qt.ItemDataRole.UserRole, boq['id'])
                    boq_list.addItem(list_item)

            def search_boqs():
                """Search BoQs based on input"""
                search_term = search_input.text().strip()
                if not search_term:
                    populate_list(all_boqs)
                else:
                    try:
                        # Search from database
                        filtered_boqs = self.db.search_cloud_boqs(search_term)
                        populate_list(filtered_boqs)
                    except:
                        # Fallback to local filtering
                        filtered_boqs = [b for b in all_boqs if search_term.lower() in b['name'].lower()]
                        populate_list(filtered_boqs)

            # Connect search input to search function
            search_input.textChanged.connect(search_boqs)

            # Initial population
            populate_list(all_boqs)

            layout.addWidget(boq_list)

            # Buttons
            button_layout = QHBoxLayout()

            load_btn = QPushButton("📥 Yüklə")
            load_btn.setDefault(True)
            delete_btn = QPushButton("🗑️ Sil")
            cancel_btn = QPushButton("❌ Ləğv Et")

            load_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)

            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
            """)

            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #9E9E9E;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #757575;
                }
            """)

            button_layout.addWidget(load_btn)
            button_layout.addWidget(delete_btn)
            button_layout.addStretch()
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)

            dialog.setLayout(layout)

            # Handle load button
            def load_selected():
                selected_items = boq_list.selectedItems()
                if not selected_items:
                    QMessageBox.warning(dialog, "Xəbərdarlıq", "BoQ seçin!")
                    return

                boq_id = selected_items[0].data(Qt.ItemDataRole.UserRole)

                # Confirm if current BoQ will be replaced
                if self.boq_items:
                    reply = QMessageBox.question(
                        dialog,
                        "Təsdiq",
                        "Mövcud BoQ məlumatları əvəz olunacaq. Davam etmək istəyirsiniz?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    if reply != QMessageBox.StandardButton.Yes:
                        return

                # Load BoQ from cloud
                boq_data = self.db.load_boq_from_cloud(boq_id)
                if boq_data:
                    self.boq_name = boq_data.get('name', 'BoQ 1')
                    self.boq_name_input.setText(self.boq_name)
                    self.next_id = boq_data.get('next_id', 1)
                    loaded_items = boq_data.get('items', [])

                    # Update prices from database for items that came from DB
                    updated_count = 0
                    for item in loaded_items:
                        if not item.get('is_custom') and item.get('product_id') and self.db:
                            try:
                                product = self.db.read_product(item['product_id'])
                                if product:
                                    old_price = item['unit_price']
                                    new_price = float(product['price']) if product.get('price') else 0

                                    item['unit_price'] = new_price
                                    item['total'] = item['quantity'] * new_price
                                    item['category'] = product.get('category', '') or ''
                                    item['source'] = product.get('mehsul_menbeyi', '') or ''
                                    item['note'] = product.get('qeyd', '') or ''

                                    if old_price != new_price:
                                        updated_count += 1
                            except:
                                pass

                    self.boq_items = loaded_items
                    self.refresh_table()

                    message = f"BoQ buluddan yükləndi!\n\n{len(loaded_items)} məhsul yükləndi."
                    if updated_count > 0:
                        message += f"\n{updated_count} məhsulun qiyməti yeniləndi."

                    QMessageBox.information(self, "Uğurlu", message)
                    dialog.accept()
                else:
                    QMessageBox.critical(dialog, "Xəta", "BoQ yüklənə bilmədi!")

            # Handle delete button
            def delete_selected():
                selected_items = boq_list.selectedItems()
                if not selected_items:
                    QMessageBox.warning(dialog, "Xəbərdarlıq", "BoQ seçin!")
                    return

                boq_id = selected_items[0].data(Qt.ItemDataRole.UserRole)
                boq_name = selected_items[0].text().split(' (')[0]

                reply = QMessageBox.question(
                    dialog,
                    "Təsdiq",
                    f"'{boq_name}' BoQ-nu buluddan silmək istədiyinizdən əminsiniz?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    if self.db.delete_cloud_boq(boq_id):
                        boq_list.takeItem(boq_list.row(selected_items[0]))
                        QMessageBox.information(dialog, "Uğurlu", "BoQ buluddan silindi!")

                        if boq_list.count() == 0:
                            QMessageBox.information(dialog, "Məlumat", "Buludda daha BoQ qalmadı.")
                            dialog.accept()
                    else:
                        QMessageBox.critical(dialog, "Xəta", "BoQ silinə bilmədi!")

            load_btn.clicked.connect(load_selected)
            delete_btn.clicked.connect(delete_selected)
            cancel_btn.clicked.connect(dialog.reject)

            # Double click to load
            boq_list.itemDoubleClicked.connect(load_selected)

            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Buluddan yükləmə xətası:\n{str(e)}")

    def combine_boqs_to_excel(self):
        """Combine multiple BoQs into a single Excel file"""
        try:
            import json
            from PyQt6.QtWidgets import QFileDialog
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # Ask user to select multiple BoQ files
            file_paths, _ = QFileDialog.getOpenFileNames(
                self,
                "BoQ Fayllarını Seçin",
                "",
                "JSON Files (*.json)"
            )

            if not file_paths or len(file_paths) < 2:
                QMessageBox.warning(self, "Xəbərdarlıq", "Ən azı 2 BoQ faylı seçin!")
                return

            # Load all BoQs
            boqs = []
            for file_path in file_paths:
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        boq_data = json.load(f)
                        boqs.append({
                            'name': boq_data.get('boq_name', os.path.basename(file_path)),
                            'items': boq_data.get('items', [])
                        })
                except Exception as e:
                    QMessageBox.warning(self, "Xəbərdarlıq", f"Fayl oxuna bilmədi: {os.path.basename(file_path)}\n{str(e)}")

            if len(boqs) < 2:
                QMessageBox.warning(self, "Xəbərdarlıq", "Ən azı 2 BoQ uğurla yüklənməlidir!")
                return

            # Create a unified list of all unique items (by name)
            all_items_dict = {}
            for boq in boqs:
                for item in boq['items']:
                    item_key = item['name']
                    if item_key not in all_items_dict:
                        all_items_dict[item_key] = {
                            'name': item['name'],
                            'unit': item.get('unit', 'ədəd'),
                            'unit_price': item.get('unit_price', 0),
                            'category': item.get('category', ''),
                            'source': item.get('source', ''),
                            'note': item.get('note', ''),
                            'quantities': {}
                        }

            # Fill quantities for each BoQ
            for boq_idx, boq in enumerate(boqs):
                boq_name = boq['name']
                for item in boq['items']:
                    item_key = item['name']
                    all_items_dict[item_key]['quantities'][boq_name] = item.get('quantity', 0)

            # Ensure all items have entries for all BoQs (fill with 0 if missing)
            for item_data in all_items_dict.values():
                for boq in boqs:
                    if boq['name'] not in item_data['quantities']:
                        item_data['quantities'][boq['name']] = 0

            # Ask user for output file
            output_path, _ = QFileDialog.getSaveFileName(
                self,
                "Birləşdirilmiş BoQ-u Yadda Saxla",
                "Birləşdirilmiş_BoQ.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not output_path:
                return

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Combined BoQ"

            # Define styles
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF", size=11)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            num_boqs = len(boqs)
            ws.merge_cells(f'A1:{chr(65 + 5 + num_boqs)}1')
            ws['A1'] = "BİRLƏŞDİRİLMİŞ BILL OF QUANTITIES (BOQ)"
            ws['A1'].font = Font(bold=True, size=16, color="2196F3")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Headers
            headers = ["№", "Adı", "Kateqoriya", "Ölçü Vahidi", "Vahid Qiymət (AZN)"]
            for boq in boqs:
                headers.append(f"{boq['name']}\n(Miqdar)")
            headers.append("Cəmi\nMiqdar")
            headers.append("Cəmi\nQiymət (AZN)")

            col_num = 1
            for header in headers:
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                col_num += 1

            # Data rows
            row_num = 4
            for idx, (item_name, item_data) in enumerate(sorted(all_items_dict.items()), 1):
                ws.cell(row=row_num, column=1, value=idx).border = border
                ws.cell(row=row_num, column=2, value=item_data['name']).border = border
                ws.cell(row=row_num, column=3, value=item_data['category'] or 'N/A').border = border
                ws.cell(row=row_num, column=4, value=item_data['unit']).border = border
                ws.cell(row=row_num, column=5, value=item_data['unit_price']).border = border
                ws.cell(row=row_num, column=5).number_format = '0.00'

                # Quantities for each BoQ
                col = 6
                first_qty_col = chr(65 + 5)  # F
                for boq in boqs:
                    qty = item_data['quantities'][boq['name']]
                    ws.cell(row=row_num, column=col, value=qty).border = border
                    ws.cell(row=row_num, column=col).number_format = '0.00'
                    col += 1

                last_qty_col = chr(65 + col - 2)  # Last BoQ column

                # Total quantity column (SUM formula)
                total_qty_col = chr(65 + col - 1)
                total_qty_cell = ws.cell(row=row_num, column=col)
                total_qty_cell.value = f"=SUM({first_qty_col}{row_num}:{last_qty_col}{row_num})"
                total_qty_cell.border = border
                total_qty_cell.number_format = '0.00'
                total_qty_cell.font = Font(bold=True)
                col += 1

                # Total price column (Total Quantity * Unit Price)
                price_col = 'E'  # Unit price column
                total_price_cell = ws.cell(row=row_num, column=col)
                total_price_cell.value = f"={total_qty_col}{row_num}*{price_col}{row_num}"
                total_price_cell.border = border
                total_price_cell.number_format = '#,##0.00'
                total_price_cell.font = Font(bold=True)

                row_num += 1

            # Add grand total row
            total_row = row_num + 1
            ws.merge_cells(f'A{total_row}:{chr(65 + 5 + num_boqs)}{total_row}')
            total_label_cell = ws[f'A{total_row}']
            total_label_cell.value = "ÜMUMİ MƏBLƏĞ:"
            total_label_cell.fill = total_fill
            total_label_cell.font = total_font
            total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_label_cell.border = border

            # Grand total formula
            total_price_col = chr(65 + 6 + num_boqs)
            grand_total_cell = ws[f'{total_price_col}{total_row}']
            grand_total_cell.value = f"=SUM({total_price_col}4:{total_price_col}{row_num - 1})"
            grand_total_cell.fill = total_fill
            grand_total_cell.font = total_font
            grand_total_cell.alignment = Alignment(horizontal="center", vertical="center")
            grand_total_cell.border = border
            grand_total_cell.number_format = '#,##0.00'

            # Adjust column widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18

            for i in range(num_boqs + 2):  # BoQ columns + Total Qty + Total Price
                col_letter = chr(70 + i)  # Start from F
                ws.column_dimensions[col_letter].width = 14

            # Save file
            wb.save(output_path)

            QMessageBox.information(
                self,
                "Uğurlu",
                f"BoQ-lar uğurla birləşdirildi!\n\n{len(boqs)} BoQ birləşdirildi\n{len(all_items_dict)} unikal məhsul\n\n{output_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"BoQ-lar birləşdirilərkən xəta:\n{str(e)}")

    def open_template_management(self):
        """Open Template Management window"""
        if not self.db:
            QMessageBox.warning(self, "Xəbərdarlıq", "Verilənlər bazasına qoşulmayıbsınız!")
            return

        dialog = TemplateManagementWindow(self, self.db, self)
        dialog.exec()


class TemplateManagementWindow(QDialog):
    """Template Management Window - Create and manage generic templates with product mapping"""

    def __init__(self, parent=None, db=None, boq_window=None):
        super().__init__(parent)
        self.db = db
        self.boq_window = boq_window  # Reference to BoQWindow for loading items
        self.current_template_id = None
        self.template_items = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("📋 Şablon İdarəetməsi")
        self.setMinimumSize(900, 600)

        main_layout = QHBoxLayout()

        # Left panel - Template list
        left_panel = QVBoxLayout()

        templates_label = QLabel("Şablonlar")
        templates_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 5px;")
        left_panel.addWidget(templates_label)

        self.template_list = QTableWidget()
        self.template_list.setColumnCount(2)
        self.template_list.setHorizontalHeaderLabels(["Şablon Adı", "Qeyd Sayı"])
        self.template_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.template_list.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.template_list.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.template_list.clicked.connect(self.on_template_selected)
        self.template_list.setMaximumWidth(300)

        header = self.template_list.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)

        left_panel.addWidget(self.template_list)

        # Template list buttons
        template_btn_layout = QHBoxLayout()

        self.new_template_btn = QPushButton("➕ Yeni")
        self.new_template_btn.clicked.connect(self.create_new_template)
        self.new_template_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 6px 12px; border: none; border-radius: 4px; font-weight: bold;")

        self.delete_template_btn = QPushButton("🗑️ Sil")
        self.delete_template_btn.clicked.connect(self.delete_template)
        self.delete_template_btn.setStyleSheet("background-color: #f44336; color: white; padding: 6px 12px; border: none; border-radius: 4px;")

        template_btn_layout.addWidget(self.new_template_btn)
        template_btn_layout.addWidget(self.delete_template_btn)
        left_panel.addLayout(template_btn_layout)

        main_layout.addLayout(left_panel)

        # Right panel - Template items
        right_panel = QVBoxLayout()

        # Template name input
        name_layout = QHBoxLayout()
        name_label = QLabel("Şablon Adı:")
        name_label.setStyleSheet("font-weight: bold;")
        self.template_name_input = QLineEdit()
        self.template_name_input.setPlaceholderText("Şablon adını daxil edin")
        self.template_name_input.setStyleSheet("padding: 8px; font-size: 14px;")
        name_layout.addWidget(name_label)
        name_layout.addWidget(self.template_name_input)
        right_panel.addLayout(name_layout)

        # Items table
        items_label = QLabel("Şablon Qeydləri")
        items_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 5px;")
        right_panel.addWidget(items_label)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(4)
        self.items_table.setHorizontalHeaderLabels(["Generik Ad", "Ölçü Vahidi", "Defolt Qiymət", "Tip"])
        self.items_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        items_header = self.items_table.horizontalHeader()
        items_header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        items_header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        items_header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        items_header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)

        right_panel.addWidget(self.items_table)

        # Item buttons
        item_btn_layout = QHBoxLayout()

        self.add_generic_btn = QPushButton("➕ Generik Qeyd")
        self.add_generic_btn.clicked.connect(self.add_generic_item)
        self.add_generic_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 8px 16px; border: none; border-radius: 4px; font-weight: bold;")
        self.add_generic_btn.setToolTip("Sərbəst generik qeyd əlavə et (məs: 'AC açar')")

        self.add_from_db_btn = QPushButton("📦 DB-dən Qeyd")
        self.add_from_db_btn.clicked.connect(self.add_item_from_db)
        self.add_from_db_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 8px 16px; border: none; border-radius: 4px; font-weight: bold;")
        self.add_from_db_btn.setToolTip("Verilənlər bazasından məhsul seç")

        self.edit_item_btn = QPushButton("✏️ Redaktə")
        self.edit_item_btn.clicked.connect(self.edit_item)
        self.edit_item_btn.setStyleSheet("background-color: #9C27B0; color: white; padding: 8px 16px; border: none; border-radius: 4px;")

        self.delete_item_btn = QPushButton("🗑️ Sil")
        self.delete_item_btn.clicked.connect(self.delete_item)
        self.delete_item_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 16px; border: none; border-radius: 4px;")

        item_btn_layout.addWidget(self.add_generic_btn)
        item_btn_layout.addWidget(self.add_from_db_btn)
        item_btn_layout.addWidget(self.edit_item_btn)
        item_btn_layout.addWidget(self.delete_item_btn)
        item_btn_layout.addStretch()

        right_panel.addLayout(item_btn_layout)

        # Save and Load to BoQ buttons
        action_btn_layout = QHBoxLayout()

        self.save_template_btn = QPushButton("💾 Şablonu Saxla")
        self.save_template_btn.clicked.connect(self.save_template)
        self.save_template_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px 20px; border: none; border-radius: 4px; font-weight: bold; font-size: 14px;")

        self.load_to_boq_btn = QPushButton("📂 BoQ-a Yüklə")
        self.load_to_boq_btn.clicked.connect(self.load_to_boq)
        self.load_to_boq_btn.setStyleSheet("background-color: #673AB7; color: white; padding: 10px 20px; border: none; border-radius: 4px; font-weight: bold; font-size: 14px;")

        action_btn_layout.addStretch()
        action_btn_layout.addWidget(self.save_template_btn)
        action_btn_layout.addWidget(self.load_to_boq_btn)

        right_panel.addLayout(action_btn_layout)

        main_layout.addLayout(right_panel)

        self.setLayout(main_layout)

        # Load existing templates
        self.refresh_template_list()

    def refresh_template_list(self):
        """Refresh the template list from database"""
        self.template_list.setRowCount(0)

        try:
            templates = self.db.get_all_templates()
            for template in templates:
                row = self.template_list.rowCount()
                self.template_list.insertRow(row)

                name_item = QTableWidgetItem(template['name'])
                name_item.setData(Qt.ItemDataRole.UserRole, template['id'])
                self.template_list.setItem(row, 0, name_item)

                item_count = len(template.get('items', []))
                self.template_list.setItem(row, 1, QTableWidgetItem(str(item_count)))
        except Exception as e:
            print(f"Error loading templates: {e}")

    def on_template_selected(self):
        """Load selected template into editor"""
        selected_row = self.template_list.currentRow()
        if selected_row < 0:
            return

        template_id = self.template_list.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)
        template_name = self.template_list.item(selected_row, 0).text()

        try:
            template = self.db.load_template(template_id)
            if template:
                self.current_template_id = template_id
                self.template_name_input.setText(template_name)
                self.template_items = template.get('items', [])
                self.refresh_items_table()
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Şablon yüklənərkən xəta: {str(e)}")

    def refresh_items_table(self):
        """Refresh the items table"""
        self.items_table.setRowCount(0)

        for item in self.template_items:
            row = self.items_table.rowCount()
            self.items_table.insertRow(row)

            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('generic_name', item.get('name', ''))))
            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('unit', '')))
            self.items_table.setItem(row, 2, QTableWidgetItem(f"{item.get('default_price', 0):.2f}"))

            # Type: Generic or DB-linked
            item_type = "DB" if item.get('product_id') else "Generik"
            self.items_table.setItem(row, 3, QTableWidgetItem(item_type))

    def create_new_template(self):
        """Create a new empty template"""
        self.current_template_id = None
        self.template_name_input.setText("")
        self.template_name_input.setPlaceholderText("Yeni şablon adı")
        self.template_items = []
        self.refresh_items_table()

    def delete_template(self):
        """Delete selected template"""
        selected_row = self.template_list.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Silmək üçün şablon seçin!")
            return

        template_name = self.template_list.item(selected_row, 0).text()
        template_id = self.template_list.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)

        reply = QMessageBox.question(
            self, "Təsdiq",
            f"'{template_name}' şablonunu silmək istədiyinizdən əminsiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.delete_template(template_id)
                self.refresh_template_list()
                if self.current_template_id == template_id:
                    self.create_new_template()
                QMessageBox.information(self, "Uğurlu", "Şablon silindi!")
            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Şablon silinərkən xəta: {str(e)}")

    def add_generic_item(self):
        """Add a generic template item"""
        dialog = TemplateItemDialog(self, mode="generic")
        if dialog.exec():
            item_data = dialog.get_data()
            self.template_items.append(item_data)
            self.refresh_items_table()

    def add_item_from_db(self):
        """Add item from database as template item"""
        dialog = TemplateItemDialog(self, mode="from_db", db=self.db)
        if dialog.exec():
            item_data = dialog.get_data()
            self.template_items.append(item_data)
            self.refresh_items_table()

    def edit_item(self):
        """Edit selected item"""
        selected_row = self.items_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Redaktə etmək üçün qeyd seçin!")
            return

        item = self.template_items[selected_row]
        mode = "from_db" if item.get('product_id') else "generic"
        dialog = TemplateItemDialog(self, mode=mode, db=self.db, item=item)
        if dialog.exec():
            item_data = dialog.get_data()
            self.template_items[selected_row] = item_data
            self.refresh_items_table()

    def delete_item(self):
        """Delete selected item"""
        selected_row = self.items_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Silmək üçün qeyd seçin!")
            return

        del self.template_items[selected_row]
        self.refresh_items_table()

    def save_template(self):
        """Save current template"""
        template_name = self.template_name_input.text().strip()
        if not template_name:
            QMessageBox.warning(self, "Xəbərdarlıq", "Şablon adı boş ola bilməz!")
            return

        if not self.template_items:
            QMessageBox.warning(self, "Xəbərdarlıq", "Şablona ən azı bir qeyd əlavə edin!")
            return

        try:
            # Convert items format for saving
            items_to_save = []
            for item in self.template_items:
                items_to_save.append({
                    'generic_name': item.get('generic_name', ''),
                    'name': item.get('name', item.get('generic_name', '')),
                    'unit': item.get('unit', ''),
                    'default_price': item.get('default_price', 0),
                    'product_id': item.get('product_id'),
                    'category': item.get('category', ''),
                    'is_generic': not bool(item.get('product_id'))
                })

            self.db.save_template(template_name, items_to_save)
            self.refresh_template_list()
            QMessageBox.information(self, "Uğurlu", f"Şablon '{template_name}' olaraq saxlanıldı!")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Şablon saxlanılarkən xəta: {str(e)}")

    def load_to_boq(self):
        """Load template items to BoQ with product selection for generic items"""
        if not self.template_items:
            QMessageBox.warning(self, "Xəbərdarlıq", "Şablonda heç bir qeyd yoxdur!")
            return

        if not self.boq_window:
            QMessageBox.warning(self, "Xəbərdarlıq", "BoQ pəncərəsi tapılmadı!")
            return

        # Ask about loading mode
        reply = QMessageBox.question(
            self, "Yükləmə Rejimi",
            "Mövcud BoQ qeydlərini əvəz etmək istəyirsiniz?\n\n'Bəli' - Əvəz et\n'Xeyr' - Mövcud qeydlərə əlavə et",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
        )

        if reply == QMessageBox.StandardButton.Cancel:
            return

        replace_mode = (reply == QMessageBox.StandardButton.Yes)

        if replace_mode:
            self.boq_window.boq_items = []
            self.boq_window.next_id = 1

        # Process each template item
        items_added = 0
        for template_item in self.template_items:
            if template_item.get('is_generic') or not template_item.get('product_id'):
                # Generic item - show product selection dialog
                selected_product = self.select_product_for_generic(template_item)
                if selected_product:
                    new_item = self.create_boq_item_from_selection(template_item, selected_product)
                    self.boq_window.boq_items.append(new_item)
                    self.boq_window.next_id += 1
                    items_added += 1
            else:
                # DB-linked item - get current data from DB
                product = self.db.read_product(template_item.get('product_id'))
                if product:
                    new_item = {
                        'id': self.boq_window.next_id,
                        'name': product['mehsulun_adi'],
                        'quantity': 1,
                        'unit': product.get('olcu_vahidi', '') or 'ədəd',
                        'unit_price': float(product['price']) if product.get('price') else 0,
                        'total': float(product['price']) if product.get('price') else 0,
                        'margin_percent': 0,
                        'category': product.get('category', ''),
                        'source': product.get('mehsul_menbeyi', ''),
                        'note': '',
                        'is_custom': False,
                        'product_id': template_item.get('product_id')
                    }
                    self.boq_window.boq_items.append(new_item)
                    self.boq_window.next_id += 1
                    items_added += 1

        self.boq_window.refresh_table()
        QMessageBox.information(self, "Uğurlu", f"{items_added} qeyd BoQ-a əlavə edildi!")
        self.accept()

    def select_product_for_generic(self, template_item):
        """Show dialog to select a product for a generic template item"""
        dialog = ProductSelectionDialog(
            self,
            self.db,
            template_item.get('generic_name', ''),
            template_item.get('category', '')
        )
        if dialog.exec():
            return dialog.get_selected_product()
        return None

    def create_boq_item_from_selection(self, template_item, product):
        """Create a BoQ item from template item and selected product"""
        return {
            'id': self.boq_window.next_id,
            'name': product['mehsulun_adi'],
            'quantity': 1,
            'unit': product.get('olcu_vahidi', '') or template_item.get('unit', 'ədəd'),
            'unit_price': float(product['price']) if product.get('price') else template_item.get('default_price', 0),
            'total': float(product['price']) if product.get('price') else template_item.get('default_price', 0),
            'margin_percent': 0,
            'category': product.get('category', ''),
            'source': product.get('mehsul_menbeyi', ''),
            'note': f"Şablondan: {template_item.get('generic_name', '')}",
            'is_custom': False,
            'product_id': str(product['_id'])
        }


class TemplateItemDialog(QDialog):
    """Dialog for adding/editing template items"""

    def __init__(self, parent=None, mode="generic", db=None, item=None):
        super().__init__(parent)
        self.mode = mode
        self.db = db
        self.item = item
        self.selected_product = None
        self.init_ui()

    def init_ui(self):
        if self.mode == "generic":
            self.setWindowTitle("Generik Qeyd Əlavə Et")
        else:
            self.setWindowTitle("DB-dən Qeyd Əlavə Et")

        self.setMinimumWidth(400)
        layout = QFormLayout()

        # Generic name
        self.generic_name_input = QLineEdit()
        self.generic_name_input.setPlaceholderText("Məs: AC açar, Kabel 2.5mm², Lampa")
        layout.addRow("Generik Ad:", self.generic_name_input)

        # Category (for filtering when loading)
        self.category_input = QLineEdit()
        self.category_input.setPlaceholderText("Kateqoriya (filtrasiya üçün)")
        layout.addRow("Kateqoriya:", self.category_input)

        # Unit
        self.unit_input = QLineEdit()
        self.unit_input.setPlaceholderText("ədəd, m, kg və s.")
        layout.addRow("Ölçü Vahidi:", self.unit_input)

        # Default price
        self.price_input = QDoubleSpinBox()
        self.price_input.setRange(0, 999999.99)
        self.price_input.setDecimals(2)
        self.price_input.setSuffix(" AZN")
        layout.addRow("Defolt Qiymət:", self.price_input)

        # Product selection (for DB mode)
        if self.mode == "from_db":
            self.product_id_input = QLineEdit()
            self.product_id_input.setPlaceholderText("Məhsul ID-ni daxil edin")
            layout.addRow("Məhsul ID:", self.product_id_input)

            self.load_product_btn = QPushButton("📥 Məhsul Yüklə")
            self.load_product_btn.clicked.connect(self.load_product_info)
            self.load_product_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 8px; border: none; border-radius: 4px;")
            layout.addRow(self.load_product_btn)

        # Fill if editing
        if self.item:
            self.generic_name_input.setText(self.item.get('generic_name', self.item.get('name', '')))
            self.category_input.setText(self.item.get('category', ''))
            self.unit_input.setText(self.item.get('unit', ''))
            self.price_input.setValue(self.item.get('default_price', 0))
            if self.mode == "from_db" and self.item.get('product_id'):
                self.product_id_input.setText(self.item.get('product_id'))

        # Buttons
        button_layout = QHBoxLayout()
        save_btn = QPushButton("💾 Saxla")
        save_btn.clicked.connect(self.accept)
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 16px; border: none; border-radius: 4px; font-weight: bold;")

        cancel_btn = QPushButton("❌ Ləğv Et")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 16px; border: none; border-radius: 4px;")

        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addRow(button_layout)

        self.setLayout(layout)

    def load_product_info(self):
        """Load product info from database"""
        if not self.db:
            return

        product_id = self.product_id_input.text().strip()
        if not product_id:
            QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul ID-ni daxil edin!")
            return

        try:
            product = self.db.read_product(product_id)
            if product:
                self.selected_product = product
                self.generic_name_input.setText(product['mehsulun_adi'])
                self.category_input.setText(product.get('category', ''))
                self.unit_input.setText(product.get('olcu_vahidi', '') or 'ədəd')
                self.price_input.setValue(float(product['price']) if product.get('price') else 0)
                QMessageBox.information(self, "Uğurlu", "Məhsul məlumatı yükləndi!")
            else:
                QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul tapılmadı!")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Xəta: {str(e)}")

    def get_data(self):
        """Get item data"""
        data = {
            'generic_name': self.generic_name_input.text().strip(),
            'name': self.generic_name_input.text().strip(),
            'category': self.category_input.text().strip(),
            'unit': self.unit_input.text().strip(),
            'default_price': self.price_input.value(),
            'is_generic': self.mode == "generic"
        }

        if self.mode == "from_db" and hasattr(self, 'product_id_input'):
            data['product_id'] = self.product_id_input.text().strip() or None

        return data

    def accept(self):
        if not self.generic_name_input.text().strip():
            QMessageBox.warning(self, "Xəbərdarlıq", "Generik ad boş ola bilməz!")
            return
        super().accept()


class ProductSelectionDialog(QDialog):
    """Dialog for selecting a product when loading a generic template item"""

    def __init__(self, parent=None, db=None, generic_name="", category=""):
        super().__init__(parent)
        self.db = db
        self.generic_name = generic_name
        self.category = category
        self.selected_product = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"Məhsul Seç: {self.generic_name}")
        self.setMinimumSize(700, 500)

        layout = QVBoxLayout()

        # Info label
        info_label = QLabel(f"'{self.generic_name}' üçün məhsul seçin:")
        info_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 5px;")
        layout.addWidget(info_label)

        # Search box
        search_layout = QHBoxLayout()
        search_label = QLabel("Axtar:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Məhsul adı ilə axtar...")
        self.search_input.setText(self.generic_name)  # Pre-fill with generic name
        self.search_input.textChanged.connect(self.search_products)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)

        # Products table
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(4)
        self.products_table.setHorizontalHeaderLabels(["ID", "Məhsul Adı", "Kateqoriya", "Qiymət"])
        self.products_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.products_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.products_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.products_table.doubleClicked.connect(self.accept)

        header = self.products_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)

        layout.addWidget(self.products_table)

        # Buttons
        button_layout = QHBoxLayout()

        select_btn = QPushButton("✅ Seç")
        select_btn.clicked.connect(self.accept)
        select_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 16px; border: none; border-radius: 4px; font-weight: bold;")

        skip_btn = QPushButton("⏭️ Keç")
        skip_btn.clicked.connect(self.reject)
        skip_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
        skip_btn.setToolTip("Bu qeydi BoQ-a əlavə etmə")

        button_layout.addStretch()
        button_layout.addWidget(select_btn)
        button_layout.addWidget(skip_btn)
        layout.addLayout(button_layout)

        self.setLayout(layout)

        # Initial search
        self.search_products()

    def search_products(self):
        """Search products by name"""
        self.products_table.setRowCount(0)
        search_text = self.search_input.text().strip()

        try:
            # Use the db's search method if available
            if hasattr(self.db, 'search_products'):
                products = self.db.search_products(search_text if search_text else None)
            else:
                products = self.db.read_all_products()
                if search_text:
                    search_lower = search_text.lower()
                    products = [p for p in products if search_lower in p.get('mehsulun_adi', '').lower()]

            for product in products[:100]:  # Limit to 100 results
                row = self.products_table.rowCount()
                self.products_table.insertRow(row)

                id_item = QTableWidgetItem(str(product['_id']))
                id_item.setData(Qt.ItemDataRole.UserRole, product)
                self.products_table.setItem(row, 0, id_item)

                self.products_table.setItem(row, 1, QTableWidgetItem(product.get('mehsulun_adi', '')))
                self.products_table.setItem(row, 2, QTableWidgetItem(product.get('category', '')))

                price = float(product['price']) if product.get('price') else 0
                self.products_table.setItem(row, 3, QTableWidgetItem(f"{price:.2f} AZN"))

        except Exception as e:
            print(f"Search error: {e}")

    def get_selected_product(self):
        """Get the selected product"""
        return self.selected_product

    def accept(self):
        selected_row = self.products_table.currentRow()
        if selected_row >= 0:
            self.selected_product = self.products_table.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)
            super().accept()
        else:
            QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul seçin!")


class ProjectWindow(QMainWindow):
    """Project Management Window"""

    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.parent_window = parent
        self.db = db
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("📁 Layihə İdarəetməsi")
        self.setGeometry(200, 200, 900, 600)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()

        # Title
        title = QLabel("Layihə İdarəetməsi")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #2196F3; padding: 10px;")
        main_layout.addWidget(title)

        # Projects table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Layihə Adı", "Təsvir", "Status", "BoQ Sayı", "Yenilənmə"])
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        self.table.cellDoubleClicked.connect(self.view_project_details)
        main_layout.addWidget(self.table)

        # Buttons
        button_layout = QHBoxLayout()

        self.new_btn = QPushButton("➕ Yeni Layihə")
        self.new_btn.clicked.connect(self.create_project)
        self.new_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)

        self.edit_btn = QPushButton("✏️ Redaktə Et")
        self.edit_btn.clicked.connect(self.edit_project)
        self.edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #0b7dda; }
        """)

        self.delete_btn = QPushButton("🗑️ Sil")
        self.delete_btn.clicked.connect(self.delete_project)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #da190b; }
        """)

        self.view_boqs_btn = QPushButton("📋 BoQ-ları Gör")
        self.view_boqs_btn.clicked.connect(self.view_project_boqs)
        self.view_boqs_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #7B1FA2; }
        """)

        self.add_boq_btn = QPushButton("➕ BoQ Əlavə Et")
        self.add_boq_btn.clicked.connect(self.add_boq_to_project)
        self.add_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #e68900; }
        """)

        self.refresh_btn = QPushButton("🔄 Yenilə")
        self.refresh_btn.clicked.connect(self.load_projects)
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #546E7A; }
        """)

        button_layout.addWidget(self.new_btn)
        button_layout.addWidget(self.edit_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.view_boqs_btn)
        button_layout.addWidget(self.add_boq_btn)
        button_layout.addWidget(self.refresh_btn)
        button_layout.addStretch()

        main_layout.addLayout(button_layout)

        # Summary label
        self.summary_label = QLabel("Layihələr yüklənir...")
        self.summary_label.setStyleSheet("color: #666; font-style: italic; padding: 5px;")
        main_layout.addWidget(self.summary_label)

        central_widget.setLayout(main_layout)

        self.setStyleSheet("""
            QMainWindow { background-color: #f5f5f5; }
            QTableWidget { background-color: white; border: 1px solid #ddd; border-radius: 5px; }
        """)

        self.load_projects()

    def load_projects(self):
        """Load all projects into the table"""
        if not self.db:
            return

        try:
            projects = self.db.get_all_projects()
            self.table.setRowCount(0)

            for project in projects:
                row = self.table.rowCount()
                self.table.insertRow(row)

                name_item = QTableWidgetItem(project['name'])
                name_item.setData(Qt.ItemDataRole.UserRole, project['id'])
                self.table.setItem(row, 0, name_item)

                self.table.setItem(row, 1, QTableWidgetItem(project.get('description', '')))
                self.table.setItem(row, 2, QTableWidgetItem(project.get('status', 'Aktiv')))
                self.table.setItem(row, 3, QTableWidgetItem(str(len(project.get('boq_ids', [])))))

                updated_at = project.get('updated_at')
                if updated_at and hasattr(updated_at, 'astimezone'):
                    date_str = updated_at.astimezone().strftime("%d.%m.%Y %H:%M")
                else:
                    date_str = "N/A"
                self.table.setItem(row, 4, QTableWidgetItem(date_str))

            self.summary_label.setText(f"Cəmi {len(projects)} layihə tapıldı")

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Layihələr yüklənərkən xəta:\n{str(e)}")

    def create_project(self):
        """Create a new project"""
        if not self.db:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Yeni Layihə")
        dialog.setMinimumWidth(400)

        layout = QFormLayout()

        name_input = QLineEdit()
        name_input.setStyleSheet("padding: 8px;")
        layout.addRow("Layihə Adı:", name_input)

        desc_input = QTextEdit()
        desc_input.setMaximumHeight(100)
        layout.addRow("Təsvir:", desc_input)

        from PyQt6.QtWidgets import QComboBox
        status_combo = QComboBox()
        status_combo.addItems(["Aktiv", "Gözləmədə", "Tamamlandı", "Ləğv edildi"])
        layout.addRow("Status:", status_combo)

        button_layout = QHBoxLayout()
        save_btn = QPushButton("💾 Saxla")
        cancel_btn = QPushButton("❌ Ləğv Et")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
        cancel_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
        save_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addRow(button_layout)

        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_input.text().strip()
            if not name:
                QMessageBox.warning(self, "Xəbərdarlıq", "Layihə adı boş ola bilməz!")
                return

            try:
                self.db.create_project(name, desc_input.toPlainText(), status_combo.currentText())
                self.load_projects()
                QMessageBox.information(self, "Uğurlu", f"Layihə '{name}' yaradıldı!")
            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Layihə yaradılarkən xəta:\n{str(e)}")

    def edit_project(self):
        """Edit selected project"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Redaktə etmək üçün layihə seçin!")
            return

        project_id = self.table.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)
        project = self.db.get_project(project_id)

        if not project:
            QMessageBox.warning(self, "Xəbərdarlıq", "Layihə tapılmadı!")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Layihəni Redaktə Et")
        dialog.setMinimumWidth(400)

        layout = QFormLayout()

        name_input = QLineEdit(project['name'])
        name_input.setStyleSheet("padding: 8px;")
        layout.addRow("Layihə Adı:", name_input)

        desc_input = QTextEdit()
        desc_input.setPlainText(project.get('description', ''))
        desc_input.setMaximumHeight(100)
        layout.addRow("Təsvir:", desc_input)

        from PyQt6.QtWidgets import QComboBox
        status_combo = QComboBox()
        status_combo.addItems(["Aktiv", "Gözləmədə", "Tamamlandı", "Ləğv edildi"])
        current_status = project.get('status', 'Aktiv')
        index = status_combo.findText(current_status)
        if index >= 0:
            status_combo.setCurrentIndex(index)
        layout.addRow("Status:", status_combo)

        button_layout = QHBoxLayout()
        save_btn = QPushButton("💾 Saxla")
        cancel_btn = QPushButton("❌ Ləğv Et")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
        cancel_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
        save_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addRow(button_layout)

        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_input.text().strip()
            if not name:
                QMessageBox.warning(self, "Xəbərdarlıq", "Layihə adı boş ola bilməz!")
                return

            try:
                self.db.update_project(project_id, name, desc_input.toPlainText(), status_combo.currentText())
                self.load_projects()
                QMessageBox.information(self, "Uğurlu", "Layihə yeniləndi!")
            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Layihə yenilənərkən xəta:\n{str(e)}")

    def delete_project(self):
        """Delete selected project"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Silmək üçün layihə seçin!")
            return

        project_name = self.table.item(selected_row, 0).text()
        project_id = self.table.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)

        reply = QMessageBox.question(
            self, "Təsdiq",
            f"'{project_name}' layihəsini silmək istədiyinizdən əminsiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.delete_project(project_id)
                self.load_projects()
                QMessageBox.information(self, "Uğurlu", "Layihə silindi!")
            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Layihə silinərkən xəta:\n{str(e)}")

    def view_project_details(self, row, column):
        """View project details on double-click"""
        self.view_project_boqs()

    def view_project_boqs(self):
        """View BoQs in selected project"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Layihə seçin!")
            return

        project_id = self.table.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)
        project_name = self.table.item(selected_row, 0).text()
        project = self.db.get_project(project_id)

        if not project:
            return

        boq_ids = project.get('boq_ids', [])

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Layihə BoQ-ları: {project_name}")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(400)

        layout = QVBoxLayout()

        # Info label
        layout.addWidget(QLabel(f"Layihədəki BoQ-lar ({len(boq_ids)}):"))

        # BoQ table
        boq_table = QTableWidget()
        boq_table.setColumnCount(3)
        boq_table.setHorizontalHeaderLabels(["BoQ Adı", "Qeyd Sayı", "Ümumi Məbləğ"])
        boq_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        boq_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        header = boq_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        total_project_amount = 0

        for boq_id in boq_ids:
            try:
                boq = self.db.load_boq_from_cloud(boq_id)
                if boq:
                    row = boq_table.rowCount()
                    boq_table.insertRow(row)

                    name_item = QTableWidgetItem(boq['name'])
                    name_item.setData(Qt.ItemDataRole.UserRole, boq_id)
                    boq_table.setItem(row, 0, name_item)

                    items = boq.get('items', [])
                    boq_table.setItem(row, 1, QTableWidgetItem(str(len(items))))

                    boq_total = sum(item.get('total', 0) for item in items)
                    total_project_amount += boq_total
                    boq_table.setItem(row, 2, QTableWidgetItem(f"{boq_total:.2f} AZN"))
            except:
                pass

        layout.addWidget(boq_table)

        # Total
        total_label = QLabel(f"Layihənin Ümumi Məbləği: {total_project_amount:.2f} AZN")
        total_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #4CAF50; padding: 10px;")
        layout.addWidget(total_label)

        # Remove button
        button_layout = QHBoxLayout()

        remove_btn = QPushButton("🗑️ Seçilmiş BoQ-u Çıxar")
        remove_btn.setStyleSheet("background-color: #f44336; color: white; padding: 8px 16px; border: none; border-radius: 4px;")

        def remove_selected():
            sel_row = boq_table.currentRow()
            if sel_row >= 0:
                boq_id_to_remove = boq_table.item(sel_row, 0).data(Qt.ItemDataRole.UserRole)
                try:
                    self.db.remove_boq_from_project(project_id, boq_id_to_remove)
                    boq_table.removeRow(sel_row)
                    self.load_projects()
                    QMessageBox.information(dialog, "Uğurlu", "BoQ layihədən çıxarıldı!")
                except Exception as e:
                    QMessageBox.critical(dialog, "Xəta", str(e))

        remove_btn.clicked.connect(remove_selected)
        button_layout.addWidget(remove_btn)

        close_btn = QPushButton("❌ Bağla")
        close_btn.setStyleSheet("background-color: #9E9E9E; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
        close_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        dialog.setLayout(layout)
        dialog.exec()

    def add_boq_to_project(self):
        """Add a cloud BoQ to selected project"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Əvvəlcə layihə seçin!")
            return

        project_id = self.table.item(selected_row, 0).data(Qt.ItemDataRole.UserRole)
        project_name = self.table.item(selected_row, 0).text()
        project = self.db.get_project(project_id)

        if not project:
            return

        try:
            all_boqs = self.db.get_all_cloud_boqs()
            existing_boq_ids = project.get('boq_ids', [])

            # Filter out already added BoQs
            available_boqs = [b for b in all_boqs if b['id'] not in existing_boq_ids]

            if not available_boqs:
                QMessageBox.information(self, "Məlumat", "Əlavə etmək üçün BoQ yoxdur. Bütün BoQ-lar artıq layihədədir.")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle(f"BoQ Əlavə Et: {project_name}")
            dialog.setMinimumWidth(500)
            dialog.setMinimumHeight(350)

            layout = QVBoxLayout()
            layout.addWidget(QLabel("Əlavə etmək istədiyiniz BoQ-u seçin:"))

            boq_table = QTableWidget()
            boq_table.setColumnCount(3)
            boq_table.setHorizontalHeaderLabels(["BoQ Adı", "Qeyd Sayı", "Yenilənmə"])
            boq_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            boq_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
            boq_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

            header = boq_table.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

            for boq in available_boqs:
                row = boq_table.rowCount()
                boq_table.insertRow(row)

                name_item = QTableWidgetItem(boq['name'])
                name_item.setData(Qt.ItemDataRole.UserRole, boq['id'])
                boq_table.setItem(row, 0, name_item)

                boq_table.setItem(row, 1, QTableWidgetItem(str(len(boq.get('items', [])))))

                updated_at = boq.get('updated_at')
                if updated_at and hasattr(updated_at, 'astimezone'):
                    date_str = updated_at.astimezone().strftime("%d.%m.%Y")
                else:
                    date_str = "N/A"
                boq_table.setItem(row, 2, QTableWidgetItem(date_str))

            layout.addWidget(boq_table)

            button_layout = QHBoxLayout()
            add_btn = QPushButton("➕ Əlavə Et")
            cancel_btn = QPushButton("❌ Ləğv Et")
            add_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px 16px; border: none; border-radius: 4px;")
            cancel_btn.setStyleSheet("background-color: #9E9E9E; color: white; padding: 8px 16px; border: none; border-radius: 4px;")

            selected_boq = [None]

            def on_add():
                sel_row = boq_table.currentRow()
                if sel_row >= 0:
                    selected_boq[0] = boq_table.item(sel_row, 0).data(Qt.ItemDataRole.UserRole)
                    dialog.accept()

            add_btn.clicked.connect(on_add)
            cancel_btn.clicked.connect(dialog.reject)
            boq_table.cellDoubleClicked.connect(lambda: on_add())

            button_layout.addWidget(add_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)

            dialog.setLayout(layout)

            if dialog.exec() == QDialog.DialogCode.Accepted and selected_boq[0]:
                self.db.add_boq_to_project(project_id, selected_boq[0])
                self.load_projects()
                QMessageBox.information(self, "Uğurlu", "BoQ layihəyə əlavə edildi!")

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"BoQ əlavə edilərkən xəta:\n{str(e)}")


class MainWindow(QMainWindow):
    """Main application window"""

    def __init__(self):
        super().__init__()
        self.db = None
        self.boq_window = None  # Single BoQ window instance
        self.project_window = None  # Single Project window instance

        # Create a timer for search debouncing
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self._perform_search)

        # Create a timer for clearing status messages
        self.status_timer = QTimer()
        self.status_timer.setSingleShot(True)
        self.status_timer.timeout.connect(self._clear_status)

        self.init_ui()
        self.show_db_config()

    def init_ui(self):
        self.setWindowTitle("🛒 Məhsul İdarəetmə Sistemi - MongoDB CRUD")
        self.setGeometry(100, 100, 1200, 700)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout
        main_layout = QVBoxLayout()

        # Title
        title = QLabel("Məhsul İdarəetmə Sistemi (MongoDB)")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setStyleSheet("color: #2196F3; padding: 10px;")
        main_layout.addWidget(title)

        # Connection status
        self.connection_label = QLabel("🔴 Verilənlər bazasına qoşulmayıb")
        self.connection_label.setStyleSheet("color: #f44336; font-weight: bold; padding: 5px;")
        self.connection_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.connection_label)

        # Search bar
        search_layout = QHBoxLayout()
        search_label = QLabel("🔍 Axtar:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Məhsul adı, kateqoriya, qeyd və ya mənbə üzrə axtar...")
        self.search_input.textChanged.connect(self.search_products)

        self.clear_search_btn = QPushButton("❌ Təmizlə")
        self.clear_search_btn.clicked.connect(self.clear_search)

        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.clear_search_btn)
        main_layout.addLayout(search_layout)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "ID", "Məhsulun Adı", "Kateqoriya", "Qiymət (AZN)", "Qiymət Dəyişdi (Gün)", "Məhsul Mənbəyi", "Ölçü Vahidi", "Qeyd"
        ])

        # Table styling
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Resize columns
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Məhsulun Adı
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Kateqoriya
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Qiymət
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Qiymət Dəyişdi (Gün)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)  # Məhsul Mənbəyi
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Ölçü Vahidi
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)  # Qeyd

        # Connect double-click to quick add to BoQ
        self.table.cellDoubleClicked.connect(self.quick_add_to_boq)

        # Enable context menu
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        main_layout.addWidget(self.table)

        # Info label
        self.info_label = QLabel("Verilənlər bazasına qoşulun...")
        self.info_label.setStyleSheet("color: #666; font-style: italic;")
        main_layout.addWidget(self.info_label)

        # Buttons
        button_layout = QHBoxLayout()

        self.add_btn = QPushButton("➕ Yeni Məhsul")
        self.add_btn.clicked.connect(self.add_product)
        self.add_btn.setEnabled(False)
        self.add_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)

        self.edit_btn = QPushButton("✏️ Redaktə Et")
        self.edit_btn.clicked.connect(self.edit_product)
        self.edit_btn.setEnabled(False)
        self.edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)

        self.delete_btn = QPushButton("🗑️ Sil")
        self.delete_btn.clicked.connect(self.delete_product)
        self.delete_btn.setEnabled(False)
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)

        self.refresh_btn = QPushButton("🔄 Yenilə")
        self.refresh_btn.clicked.connect(self.load_products)
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e68900;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)

        self.reconnect_btn = QPushButton("🔌 Yenidən Qoşul")
        self.reconnect_btn.clicked.connect(self.show_db_config)
        self.reconnect_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)

        self.add_to_boq_btn = QPushButton("➕ BoQ-ya Əlavə Et")
        self.add_to_boq_btn.clicked.connect(self.add_selected_to_boq)
        self.add_to_boq_btn.setEnabled(False)
        self.add_to_boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #673AB7;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5E35B1;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)

        self.boq_btn = QPushButton("📋 BoQ Aç")
        self.boq_btn.clicked.connect(self.open_boq_window)
        self.boq_btn.setStyleSheet("""
            QPushButton {
                background-color: #00BCD4;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #00ACC1;
            }
        """)

        self.project_btn = QPushButton("📁 Layihələr")
        self.project_btn.clicked.connect(self.open_project_window)
        self.project_btn.setStyleSheet("""
            QPushButton {
                background-color: #795548;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #6D4C41;
            }
        """)

        button_layout.addWidget(self.add_btn)
        button_layout.addWidget(self.edit_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.add_to_boq_btn)
        button_layout.addWidget(self.refresh_btn)
        button_layout.addWidget(self.reconnect_btn)
        button_layout.addWidget(self.boq_btn)
        button_layout.addWidget(self.project_btn)
        button_layout.addStretch()

        main_layout.addLayout(button_layout)

        central_widget.setLayout(main_layout)

        # Window styling
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 13px;
            }
        """)

        # Setup keyboard shortcuts
        self.setup_shortcuts()

    def setup_shortcuts(self):
        """Setup keyboard shortcuts for MainWindow"""
        # Ctrl+N: Add new product
        QShortcut(QKeySequence("Ctrl+N"), self).activated.connect(self.add_product)

        # Ctrl+E: Edit selected product
        QShortcut(QKeySequence("Ctrl+E"), self).activated.connect(self.edit_product)

        # Delete: Delete selected product(s)
        QShortcut(QKeySequence("Delete"), self).activated.connect(self.delete_product)

        # Ctrl+F: Focus search box
        QShortcut(QKeySequence("Ctrl+F"), self).activated.connect(self.focus_search)

        # Ctrl+B: Open BoQ window
        QShortcut(QKeySequence("Ctrl+B"), self).activated.connect(self.open_boq_window)

        # F5: Refresh products
        QShortcut(QKeySequence("F5"), self).activated.connect(self.load_products)

    def focus_search(self):
        """Focus the search input field"""
        self.search_input.setFocus()
        self.search_input.selectAll()

    def show_db_config(self):
        """Show database configuration dialog"""
        dialog = DatabaseConfigDialog(self)
        if dialog.exec():
            config = dialog.get_config()
            try:
                self.db = DatabaseManager(**config)
                user_display = f"{config['username']}@" if config['username'] else ""
                self.connection_label.setText(
                    f"🟢 Qoşuldu: {user_display}{config['host']}:{config['port']}/{config['database']}"
                )
                self.connection_label.setStyleSheet("color: #4CAF50; font-weight: bold; padding: 5px;")

                # Enable buttons
                self.add_btn.setEnabled(True)
                self.edit_btn.setEnabled(True)
                self.delete_btn.setEnabled(True)
                self.add_to_boq_btn.setEnabled(True)
                self.refresh_btn.setEnabled(True)

                self.load_products()

            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Verilənlər Bazası Xətası",
                    f"Verilənlər bazasına qoşula bilmədi:\n{str(e)}"
                )

    def load_products(self, preserve_status=False):
        """Load all products into table"""
        if not self.db:
            return

        try:
            products = self.db.read_all_products()
            self.populate_table(products)
            if not preserve_status:
                self.info_label.setText(f"Cəmi məhsul sayı: {len(products)}")
                self.info_label.setStyleSheet("color: #666; font-style: italic;")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Məhsullar yüklənə bilmədi:\n{str(e)}")

    def populate_table(self, products):
        """Populate table with product data"""
        self.table.setRowCount(0)

        for product in products:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            # Store the ObjectId string in the table
            self.table.setItem(row_position, 0, QTableWidgetItem(product['id']))
            self.table.setItem(row_position, 1, QTableWidgetItem(product['mehsulun_adi']))
            self.table.setItem(row_position, 2, QTableWidgetItem(
                product.get('category', '') or 'N/A'
            ))
            self.table.setItem(row_position, 3, QTableWidgetItem(
                f"{float(product['price']):.2f}" if product.get('price') else "N/A"
            ))

            # Calculate days since price last changed
            price_last_changed = product.get('price_last_changed')
            if price_last_changed:
                # Handle both timezone-aware and naive datetime
                if price_last_changed.tzinfo is None:
                    price_last_changed = price_last_changed.replace(tzinfo=timezone.utc)

                now = datetime.now(timezone.utc)
                days_since_change = (now - price_last_changed).days

                # Color code based on days
                days_item = QTableWidgetItem(str(days_since_change))
                if days_since_change > 365:
                    days_item.setForeground(QColor(211, 47, 47))  # Red for over a year
                elif days_since_change > 180:
                    days_item.setForeground(QColor(255, 152, 0))  # Orange for over 6 months
                elif days_since_change > 90:
                    days_item.setForeground(QColor(255, 193, 7))  # Yellow for over 3 months
                else:
                    days_item.setForeground(QColor(76, 175, 80))  # Green for recent

                self.table.setItem(row_position, 4, days_item)
            else:
                self.table.setItem(row_position, 4, QTableWidgetItem("N/A"))

            self.table.setItem(row_position, 5, QTableWidgetItem(
                product.get('mehsul_menbeyi', '') or 'N/A'
            ))
            self.table.setItem(row_position, 6, QTableWidgetItem(
                product.get('olcu_vahidi', '') or 'N/A'
            ))
            self.table.setItem(row_position, 7, QTableWidgetItem(
                product.get('qeyd', '') or 'N/A'
            ))

    def add_product(self):
        """Open dialog to add new product"""
        if not self.db:
            return

        # Open dialog - it handles saving and clearing internally
        dialog = ProductDialog(self, mode="add")
        dialog.exec()
        # No need to handle the result - dialog manages everything

    def edit_product(self):
        """Edit selected product"""
        if not self.db:
            return

        selected_row = self.table.currentRow()

        if selected_row < 0:
            QMessageBox.warning(self, "Xəbərdarlıq", "Zəhmət olmasa redaktə etmək üçün məhsul seçin!")
            return

        product_id = self.table.item(selected_row, 0).text()

        try:
            product = self.db.read_product(product_id)

            if not product:
                QMessageBox.critical(self, "Xəta", "Məhsul tapılmadı!")
                return

            dialog = ProductDialog(self, product=product, mode="edit")
            if dialog.exec():
                data = dialog.get_data()

                # Handle image upload
                image_id = None
                if data['image_data']:
                    # New image uploaded
                    image_id = self.db.save_image(data['image_data'], data['image_filename'])
                elif data['remove_image']:
                    # Image removed
                    image_id = ''  # Empty string signals removal

                success = self.db.update_product(
                    product_id,
                    data['mehsulun_adi'],
                    data['price'],
                    data['mehsul_menbeyi'],
                    data['qeyd'],
                    data['olcu_vahidi'],
                    data['category'],
                    image_id=image_id if image_id is not None else None
                )
                if success:
                    QMessageBox.information(self, "Uğurlu", "Məhsul uğurla yeniləndi!")
                    self.load_products()
                else:
                    QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul yenilənə bilmədi!")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Məhsul yenilənə bilmədi:\n{str(e)}")

    def delete_product(self):
        """Delete selected product(s)"""
        if not self.db:
            return

        # Get all selected rows
        selected_rows = self.table.selectionModel().selectedRows()

        if not selected_rows:
            QMessageBox.warning(self, "Xəbərdarlıq", "Zəhmət olmasa silmək üçün məhsul seçin!")
            return

        # Collect product IDs and names
        products_to_delete = []
        for index in selected_rows:
            row = index.row()
            product_id = self.table.item(row, 0).text()
            product_name = self.table.item(row, 1).text()
            products_to_delete.append((product_id, product_name))

        # Confirm deletion
        if len(products_to_delete) == 1:
            message = f"'{products_to_delete[0][1]}' məhsulunu silmək istədiyinizdən əminsiniz?"
        else:
            message = f"{len(products_to_delete)} məhsulu silmək istədiyinizdən əminsiniz?"

        reply = QMessageBox.question(
            self,
            "Təsdiq",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                deleted_count = 0
                failed_count = 0

                for product_id, product_name in products_to_delete:
                    success = self.db.delete_product(product_id)
                    if success:
                        deleted_count += 1
                    else:
                        failed_count += 1

                if deleted_count > 0:
                    QMessageBox.information(
                        self,
                        "Uğurlu",
                        f"{deleted_count} məhsul uğurla silindi!" +
                        (f"\n{failed_count} məhsul silinə bilmədi." if failed_count > 0 else "")
                    )
                    self.load_products()
                else:
                    QMessageBox.warning(self, "Xəbərdarlıq", "Heç bir məhsul silinə bilmədi!")
            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Məhsul silinə bilmədi:\n{str(e)}")

    def view_product_image(self):
        """View product image from context menu"""
        if not self.db:
            return

        selected_row = self.table.currentRow()
        if selected_row < 0:
            return

        try:
            # Get product ID from the selected row
            product_id = self.table.item(selected_row, 0).text()
            product_name = self.table.item(selected_row, 1).text()

            # Retrieve product data
            product = self.db.read_product(product_id)

            if not product:
                QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul tapılmadı!")
                return

            # Check if product has an image
            if product.get('image_id'):
                try:
                    # Retrieve image from GridFS
                    image_data = self.db.get_image(product['image_id'])

                    # Show image viewer dialog
                    viewer = ImageViewerDialog(self, image_data, product_name)
                    viewer.exec()
                except Exception as e:
                    QMessageBox.warning(self, "Xəbərdarlıq", f"Şəkil yüklənə bilmədi: {e}")
            else:
                QMessageBox.information(self, "Məlumat", f"'{product_name}' üçün şəkil yoxdur.\n\nŞəkil əlavə etmək üçün məhsulu redaktə edin.")

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Şəkil göstərilə bilmədi:\n{str(e)}")

    def show_price_history(self):
        """Show price history dialog for selected product"""
        if not self.db:
            return

        selected_row = self.table.currentRow()
        if selected_row < 0:
            return

        try:
            product_id = self.table.item(selected_row, 0).text()
            product_name = self.table.item(selected_row, 1).text()

            product = self.db.read_product(product_id)
            if not product:
                QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul tapılmadı!")
                return

            price_history = self.db.get_price_history(product_id)
            current_price = product.get('price', 0)

            dialog = PriceHistoryDialog(self, product_name, price_history, current_price)
            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Qiymət tarixi göstərilə bilmədi:\n{str(e)}")

    def show_context_menu(self, position):
        """Show context menu on right-click"""
        if not self.db:
            return

        # Get selected rows
        selected_rows = self.table.selectionModel().selectedRows()

        if not selected_rows:
            return

        # Create context menu
        menu = QMenu(self)

        # Single selection actions
        if len(selected_rows) == 1:
            edit_action = menu.addAction("✏️ Redaktə Et")
            edit_action.triggered.connect(self.edit_product)

            copy_action = menu.addAction("📋 Kopyala və Redaktə Et")
            copy_action.triggered.connect(self.copy_product)

            image_action = menu.addAction("🖼️ Şəkli Göstər")
            image_action.triggered.connect(self.view_product_image)

            price_history_action = menu.addAction("📈 Qiymət Tarixi")
            price_history_action.triggered.connect(self.show_price_history)

            menu.addSeparator()

        # Multi-selection compatible actions
        add_to_boq_action = menu.addAction(f"➕ BoQ-a Əlavə Et ({len(selected_rows)} məhsul)")
        add_to_boq_action.triggered.connect(self.add_selected_to_boq)

        menu.addSeparator()

        delete_action = menu.addAction(f"🗑️ Sil ({len(selected_rows)} məhsul)")
        delete_action.triggered.connect(self.delete_product)

        # Show menu at cursor position
        menu.exec(self.table.viewport().mapToGlobal(position))

    def copy_product(self):
        """Copy selected product and open for editing"""
        if not self.db:
            return

        selected_row = self.table.currentRow()
        if selected_row < 0:
            return

        try:
            # Get product ID and data
            product_id = self.table.item(selected_row, 0).text()
            product = self.db.read_product(product_id)

            if not product:
                QMessageBox.critical(self, "Xəta", "Məhsul tapılmadı!")
                return

            # Create a copy of the product (without _id and image_id)
            product_copy = {
                'mehsulun_adi': product['mehsulun_adi'] + " (kopya)",
                'price': product.get('price'),
                'mehsul_menbeyi': product.get('mehsul_menbeyi'),
                'qeyd': product.get('qeyd'),
                'olcu_vahidi': product.get('olcu_vahidi'),
                'category': product.get('category')
            }

            # Open dialog in "add" mode with copied data
            dialog = ProductDialog(self, product=product_copy, mode="add")
            dialog.setWindowTitle("Məhsulu Kopyala")

            if dialog.exec():
                data = dialog.get_data()

                # Handle image if copied
                image_id = None
                if data['image_data']:
                    image_id = self.db.save_image(data['image_data'], data['image_filename'])

                # Create new product
                new_product_id = self.db.create_product(
                    data['mehsulun_adi'],
                    data['price'],
                    data['mehsul_menbeyi'],
                    data['qeyd'],
                    data['olcu_vahidi'],
                    data['category'],
                    image_id=image_id
                )

                if new_product_id:
                    QMessageBox.information(self, "Uğurlu", "Məhsul kopyalandı və əlavə edildi!")
                    self.load_products()
                else:
                    QMessageBox.warning(self, "Xəbərdarlıq", "Məhsul kopyalana bilmədi!")

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Məhsul kopyalanarkən xəta:\n{str(e)}")

    def quick_add_to_boq(self, row, column):
        """Quick add product to BoQ on double-click"""
        if not self.db:
            QMessageBox.warning(self, "Xəta", "Verilənlər bazasına qoşulmayıbsınız!")
            return

        # Check if BoQ window exists
        if not self.boq_window:
            QMessageBox.warning(self, "Xəbərdarlıq", "BoQ pəncərəsi açılmayıb! Əvvəlcə 'BoQ Aç' düyməsini basın.")
            return

        try:
            product_id = self.table.item(row, 0).text()
            product = self.db.read_product(product_id)

            if not product:
                QMessageBox.warning(self, "Xəta", f"Məhsul tapılmadı: {product_id}")
                return

            # Create a pre-filled item for the dialog
            item = {
                'name': product['mehsulun_adi'],
                'quantity': 1,
                'unit': product.get('olcu_vahidi', ''),
                'unit_price': product.get('price', 0),
                'category': product.get('category', ''),
                'source': product.get('mehsul_menbeyi', ''),
                'note': product.get('qeyd', '')
            }

            # Show dialog for quantity input
            dialog = BoQItemDialog(self, self.db, item=item, mode="edit")
            dialog.setWindowTitle(f"BoQ-a Əlavə Et: {product['mehsulun_adi']}")

            if dialog.exec():
                data = dialog.get_data()
                data['id'] = self.boq_window.next_id
                self.boq_window.next_id += 1
                self.boq_window.boq_items.append(data)
                self.boq_window.refresh_table()
                self.show_status(f"'{product['mehsulun_adi']}' BoQ-a əlavə edildi", "#4CAF50")

        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Məhsul əlavə edilərkən xəta:\n{str(e)}")

    def add_selected_to_boq(self):
        """Add selected products to BoQ with sequential quantity dialogs"""
        if not self.db:
            QMessageBox.warning(self, "Xəta", "Verilənlər bazasına qoşulmayıbsınız!")
            return

        # Check if BoQ window exists
        if not self.boq_window:
            QMessageBox.warning(self, "Xəbərdarlıq", "BoQ pəncərəsi açılmayıb! Əvvəlcə 'BoQ yarat' düyməsini basın.")
            return

        # Get all selected rows
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return

        # Process each selected product
        added_count = 0
        for index in selected_rows:
            row = index.row()
            product_id = self.table.item(row, 0).text()

            try:
                # Load product from database
                product = self.db.read_product(product_id)
                if not product:
                    QMessageBox.warning(self, "Xəta", f"Məhsul tapılmadı: {product_id}")
                    continue

                # Create a pre-filled item for the dialog
                item = {
                    'name': product['mehsulun_adi'],
                    'quantity': 1,
                    'unit': product.get('olcu_vahidi', ''),
                    'unit_price': product.get('price', 0),
                    'category': product.get('category', '')
                }

                # Show dialog for quantity input
                dialog = BoQItemDialog(self, self.db, item=item, mode="edit")
                dialog.setWindowTitle(f"BoQ-a Əlavə Et: {product['mehsulun_adi']}")

                if dialog.exec():
                    # Get the data from dialog
                    data = dialog.get_data()
                    data['id'] = self.boq_window.next_id
                    self.boq_window.next_id += 1

                    # Add to BoQ
                    self.boq_window.boq_items.append(data)
                    added_count += 1
                else:
                    # User cancelled, stop processing remaining items
                    break

            except Exception as e:
                QMessageBox.critical(self, "Xəta", f"Məhsul əlavə edilərkən xəta:\n{str(e)}")
                continue

        # Refresh BoQ table and show success message
        if added_count > 0:
            self.boq_window.refresh_table()
            if added_count == 1:
                self.show_status(f"1 məhsul BoQ-a əlavə edildi", "#4CAF50")
            else:
                self.show_status(f"{added_count} məhsul BoQ-a əlavə edildi", "#4CAF50")

    def search_products(self):
        """Debounced search - waits for user to stop typing"""
        # Stop any existing timer
        self.search_timer.stop()

        # Start a new timer that will trigger search after 300ms of no typing
        self.search_timer.start(300)

    def _perform_search(self):
        """Actually perform the search (called by timer)"""
        if not self.db:
            return

        search_term = self.search_input.text().strip()

        if not search_term:
            self.load_products()
            return

        try:
            products = self.db.search_products(search_term)
            self.populate_table(products)
            self.info_label.setText(f"Axtarış nəticəsi: {len(products)} məhsul tapıldı")
        except Exception as e:
            QMessageBox.critical(self, "Xəta", f"Axtarış zamanı xəta:\n{str(e)}")

    def clear_search(self):
        """Clear search and reload all products"""
        self.search_timer.stop()  # Stop any pending search
        self.search_input.clear()
        self.load_products()

    def show_status(self, message, color="#4CAF50", duration=3000):
        """Show a temporary status message at the bottom"""
        self.info_label.setText(message)
        self.info_label.setStyleSheet(f"color: {color}; font-weight: bold; padding: 5px;")

        # Clear the status after duration (default 3 seconds)
        self.status_timer.stop()
        self.status_timer.start(duration)

    def _clear_status(self):
        """Clear status message and show product count"""
        if self.db:
            try:
                products = self.db.read_all_products()
                self.info_label.setText(f"Cəmi məhsul sayı: {len(products)}")
                self.info_label.setStyleSheet("color: #666; font-style: italic;")
            except:
                pass

    def open_boq_window(self):
        """Open the Bill of Quantities window (singleton)"""
        if self.boq_window is None or not self.boq_window.isVisible():
            self.boq_window = BoQWindow(self, self.db)
            self.boq_window.show()
        else:
            # Bring existing window to front
            self.boq_window.raise_()
            self.boq_window.activateWindow()

    def open_project_window(self):
        """Open the Project Management window (singleton)"""
        if not self.db:
            QMessageBox.warning(self, "Xəbərdarlıq", "Əvvəlcə verilənlər bazasına qoşulun!")
            return

        if self.project_window is None or not self.project_window.isVisible():
            self.project_window = ProjectWindow(self, self.db)
            self.project_window.show()
        else:
            # Bring existing window to front
            self.project_window.raise_()
            self.project_window.activateWindow()


def main():
    app = QApplication(sys.argv)

    # Set application style
    app.setStyle('Fusion')

    # Force light mode palette
    light_palette = QPalette()
    light_palette.setColor(QPalette.ColorRole.Window, QColor(245, 245, 245))
    light_palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))
    light_palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
    light_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(240, 240, 240))
    light_palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(255, 255, 220))
    light_palette.setColor(QPalette.ColorRole.ToolTipText, QColor(0, 0, 0))
    light_palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0))
    light_palette.setColor(QPalette.ColorRole.Button, QColor(240, 240, 240))
    light_palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))
    light_palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 0, 0))
    light_palette.setColor(QPalette.ColorRole.Link, QColor(0, 122, 204))
    light_palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
    light_palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))

    app.setPalette(light_palette)

    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
